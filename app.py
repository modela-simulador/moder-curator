#!/usr/bin/env python3
"""
MODÈR Product Curator v2
Web interface for curating crawled products into the MODÈR import spreadsheet.
"""

import os
import json
import io
import re
import xml.etree.ElementTree as ET
import requests
from bs4 import BeautifulSoup
from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for
from firestore_storage import (
    save_session_firestore, load_session_firestore,
    save_brands_firestore, load_brands_firestore,
    save_country_firestore, load_country_firestore,
    save_cache_firestore, load_cache_firestore,
    clear_session_firestore, clear_cache_firestore, clear_all_firestore,
    is_firestore_available, _get_db, firestore_timestamp
)
import threading
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import time
from datetime import datetime

app = Flask(__name__)
app.json.ensure_ascii = True  # Security: escapes < as \u003c in tojson, preventing XSS in <script> tags

# Secret key MUST be set via environment variable — no weak fallback
_secret = os.environ.get("SECRET_KEY", "")
if not _secret:
    import secrets
    _secret = secrets.token_hex(32)
    print("⚠️ SECRET_KEY not set — generated random key (sessions won't survive restarts)")
app.secret_key = _secret

app.config['SESSION_COOKIE_SECURE'] = True      # Required for Safari Private Browsing
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'   # Prevents cookie being dropped on cross-site
app.config['SESSION_COOKIE_HTTPONLY'] = True      # XSS protection
app.config['SESSION_PERMANENT'] = True
from datetime import timedelta
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=7)

# ─── Auth ────────────────────────────────────────────────────────────────
# Credentials from env vars. Format: "user1:hash1,user2:hash2"
# Generate hashes: python -c "from werkzeug.security import generate_password_hash; print(generate_password_hash('mypassword'))"
from werkzeug.security import check_password_hash, generate_password_hash

def _load_users():
    """Load users from CURATOR_USERS env var, fallback to defaults for dev."""
    env_users = os.environ.get("CURATOR_USERS", "")
    if env_users:
        users = {}
        for entry in env_users.split(","):
            if ":" in entry:
                name, pw_hash = entry.split(":", 1)
                users[name.strip()] = pw_hash.strip()
        if users:
            return users
    # Fallback for development — will print warning
    print("⚠️ CURATOR_USERS not set — using default credentials (set env var for production)")
    default_pw = generate_password_hash("moder2026", method="pbkdf2:sha256")
    return {
        "sebastian": default_pw,
        "antonia": default_pw,
        "sofia": default_pw,
    }

USERS = _load_users()

from functools import wraps
from flask import session as flask_session

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not flask_session.get("logged_in"):
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return decorated

def get_user_id():
    """Get current user ID from Flask session — used to isolate curation sessions"""
    return flask_session.get("username", "default")

def _safe_get_user_id():
    """Get user ID without crashing outside request context (for background threads)."""
    try:
        return flask_session.get("username", "anonymous")
    except RuntimeError:
        return "anonymous"

# CSRF note: not needed because SESSION_COOKIE_SAMESITE='Lax' prevents
# cross-origin POSTs from attaching the session cookie. All state-changing
# endpoints also require login_required. This is sufficient protection.

# ─── Config ──────────────────────────────────────────────────────────────
# Persistent disk on Render (survives redeploys), fallback to local dir
PERSISTENT_DIR = "/opt/render/project/src/data"
if os.path.isdir(PERSISTENT_DIR):
    DATA_DIR = PERSISTENT_DIR
else:
    DATA_DIR = os.path.dirname(os.path.abspath(__file__))
CRAWL_CACHE = os.path.join(DATA_DIR, "crawl_cache.json")
SESSION_FILE = os.path.join(DATA_DIR, "session.json")

DEFAULT_BRANDS = []  # Vacío — el usuario elige desde las sugeridas o agrega manualmente

# ─── Countries ──────────────────────────────────────────────────────────
COUNTRIES = {
    "CL": {"name": "Chile", "flag": "🇨🇱", "currency": "CLP"},
    "AR": {"name": "Argentina", "flag": "🇦🇷", "currency": "ARS"},
    "MX": {"name": "México", "flag": "🇲🇽", "currency": "MXN"},
    "CO": {"name": "Colombia", "flag": "🇨🇴", "currency": "COP"},
    "ES": {"name": "España", "flag": "🇪🇸", "currency": "EUR"},
}

# Marcas sugeridas por país — cada país tiene su propio catálogo
SUGGESTED_BRANDS_BY_COUNTRY = {
    "CL": [
        {"name": "CASSIOPEA", "domain": "cassiopeaofficial.com", "url": "https://shop.cassiopeaofficial.com"},
        {"name": "PARSOME", "domain": "parsome.cl", "url": "https://www.parsome.cl"},
        {"name": "ARDE", "domain": "wearearde.cl", "url": "https://wearearde.cl"},
        {"name": "LA COT", "domain": "lacotmuet.cl", "url": "https://lacotmuet.cl", "platform": "woocommerce"},
        {"name": "D.GARCÍA", "domain": "degarcia.cl", "url": "https://www.degarcia.cl"},
        {"name": "ANTONIA FLUXÁ", "domain": "antoniafluxa.cl", "url": "https://www.antoniafluxa.cl"},
        {"name": "OCHI AND CO.", "domain": "ochiandco.cl", "url": "https://www.ochiandco.cl"},
        {"name": "FRANCA E IO", "domain": "francaeio.cl", "url": "https://www.francaeio.cl", "platform": "jumpseller"},
        {"name": "ATHAR", "domain": "atharshoes.cl", "url": "https://www.atharshoes.cl"},
        {"name": "AMBAR", "domain": "tiendaambar.cl", "url": "https://www.tiendaambar.cl"},
        {"name": "ANONIMATO SHOP", "domain": "anonimato.cl", "url": "https://www.anonimato.cl"},
        {"name": "LOLITA LPK", "domain": "lpk.cl", "url": "https://www.lpk.cl"},
        {"name": "MARINA MIA", "domain": "marinamia.cl", "url": "https://www.marinamia.cl"},
        {"name": "DEBUT", "domain": "debut.cl", "url": "https://www.debut.cl"},
        {"name": "VIELLA", "domain": "viella.cl", "url": "https://www.viella.cl"},
        {"name": "LORAINE HOLMES", "domain": "loraineholmes.cl", "url": "https://www.loraineholmes.cl"},
        {"name": "CANDELARIA PÉREZ", "domain": "candelariaperez.cl", "url": "https://www.candelariaperez.cl"},
        {"name": "COCO LABEL", "domain": "cocolabel.cl", "url": "https://www.cocolabel.cl"},
        {"name": "MARIA GULDMAN", "domain": "mariaguldman.cl", "url": "https://www.mariaguldman.cl"},
        {"name": "CAROLINA FLORES", "domain": "carolinafloreshandmade.cl", "url": "https://carolinafloreshandmade.cl"},
        {"name": "ADEU.", "domain": "adeu.cl", "url": "https://www.adeu.cl"},
        {"name": "SAINTMALE", "domain": "saintmale.com", "url": "https://www.saintmale.com"},
        {"name": "CAIS.", "domain": "caiszapatos.com", "url": "https://www.caiszapatos.com"},
        {"name": "MANTO SILVESTRE", "domain": "mantosilvestre.cl", "url": "https://www.mantosilvestre.cl"},
    ],
    "AR": [],
    "MX": [],
    "CO": [],
    "ES": [],
}

# Flat list for backward compatibility
SUGGESTED_BRANDS = SUGGESTED_BRANDS_BY_COUNTRY.get("CL", [])

# ─── Country-aware file paths ───────────────────────────────────────────
COUNTRY_FILE = os.path.join(DATA_DIR, "active_country.json")

def load_active_country():
    """Per-user country from Flask session, with Firestore fallback"""
    cc = flask_session.get("active_country", "")
    if not cc:
        # Fallback: try Firestore per-user country
        uid = get_user_id()
        fs = load_country_firestore()
        if fs and isinstance(fs, dict):
            cc = fs.get(uid, "")
        elif fs and isinstance(fs, str):
            cc = fs
        if cc:
            flask_session["active_country"] = cc
    return cc

def save_active_country(country_code):
    """Per-user country in Flask session + Firestore backup"""
    flask_session["active_country"] = country_code
    try:
        uid = get_user_id()
        save_country_firestore({uid: country_code})
    except Exception:
        pass

def get_brands_file_for_country(country_code):
    """Each country gets its own active brands file"""
    if country_code:
        return os.path.join(DATA_DIR, f"active_brands_{country_code}.json")
    return os.path.join(DATA_DIR, "active_brands.json")

def get_cache_file_for_country(country_code, user_id=None):
    """Each user+country gets its own crawl cache"""
    uid = user_id or get_user_id()
    if country_code:
        return os.path.join(DATA_DIR, f"crawl_cache_{uid}_{country_code}.json")
    return CRAWL_CACHE

BRANDS_FILE = os.path.join(DATA_DIR, "active_brands.json")

def _brands_file_for_user(country_code, user_id=None):
    uid = user_id or get_user_id()
    return os.path.join(DATA_DIR, f"brands_{uid}_{country_code}.json")

def _hidden_brands_file(country_code, user_id=None):
    uid = user_id or get_user_id()
    return os.path.join(DATA_DIR, f"hidden_brands_{uid}_{country_code}.json")

def load_hidden_brands(country_code=None, user_id=None):
    """Load list of permanently hidden brand domains"""
    uid = user_id or get_user_id()
    cc = country_code or load_active_country()
    if not cc:
        return []
    # Firestore first
    db = _get_db_safe()
    if db:
        try:
            doc = db.collection("curator").document(f"hidden_{uid}_{cc}").get()
            if doc.exists:
                return doc.to_dict().get("domains", [])
        except Exception:
            pass
    # Local file fallback
    path = _hidden_brands_file(cc, uid)
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return []

def save_hidden_brands(domains, country_code=None, user_id=None):
    """Save list of permanently hidden brand domains"""
    uid = user_id or get_user_id()
    cc = country_code or load_active_country()
    if not cc:
        return
    path = _hidden_brands_file(cc, uid)
    tmp_path = path + ".tmp"
    with open(tmp_path, "w") as f:
        json.dump(domains, f, ensure_ascii=False)
        f.flush()
        os.fsync(f.fileno())
    os.replace(tmp_path, path)
    db = _get_db_safe()
    if db:
        try:
            db.collection("curator").document(f"hidden_{uid}_{cc}").set({
                "domains": domains, "updated_at": firestore_timestamp()
            })
        except Exception:
            pass

def _get_db_safe():
    """Get Firestore db, return None if unavailable"""
    try:
        return _get_db()
    except Exception:
        return None

def load_active_brands(country_code=None, user_id=None):
    """Load active brands per user+country"""
    uid = user_id or get_user_id()
    cc = country_code or load_active_country()
    if not cc:
        return DEFAULT_BRANDS
    # Firestore primero (per user+country)
    fs = load_brands_firestore(f"{uid}_{cc}")
    if fs is not None:
        return fs
    # Fallback archivo per user
    path = _brands_file_for_user(cc, uid)
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    # Legacy fallback (shared file)
    legacy = get_brands_file_for_country(cc)
    if os.path.exists(legacy):
        with open(legacy) as f:
            return json.load(f)
    return DEFAULT_BRANDS

def save_active_brands(brands, country_code=None, user_id=None):
    """Save active brands per user+country"""
    uid = user_id or get_user_id()
    cc = country_code or load_active_country()
    if not cc:
        return
    path = _brands_file_for_user(cc, uid)
    tmp_path = path + ".tmp"
    with open(tmp_path, "w") as f:
        json.dump(brands, f, ensure_ascii=False, indent=2)
        f.flush()
        os.fsync(f.fileno())
    os.replace(tmp_path, path)
    save_brands_firestore(brands, f"{uid}_{cc}")

HEADERS = {
    "User-Agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Mobile/15E148 Safari/604.1",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "es-CL,es;q=0.9,en;q=0.8",
}

# ─── Robots.txt cache ────────────────────────────────────────────────────
_robots_cache = {}  # {domain: (allowed: bool, timestamp)} — max 200 entries

def is_crawl_allowed(url, user_agent="*"):
    """Check robots.txt for a URL. Cached per domain for 1 hour."""
    from urllib.parse import urlparse
    import time as _time
    parsed = urlparse(url)
    domain = parsed.netloc
    now = _time.time()

    # Check cache (1 hour TTL)
    if domain in _robots_cache:
        allowed, ts = _robots_cache[domain]
        if now - ts < 3600:
            return allowed

    # Evict oldest entries if cache is too large
    if len(_robots_cache) > 200:
        oldest = min(_robots_cache, key=lambda k: _robots_cache[k][1])
        del _robots_cache[oldest]

    try:
        from urllib.robotparser import RobotFileParser
        robots_url = f"{parsed.scheme}://{domain}/robots.txt"
        rp = RobotFileParser()
        rp.set_url(robots_url)
        rp.read()
        allowed = rp.can_fetch("*", url)
        _robots_cache[domain] = (allowed, now)
        return allowed
    except Exception:
        _robots_cache[domain] = (True, now)  # On error, allow crawl
        return True

# ─── Session state (with in-memory cache) ────────────────────────────────

_session_cache = {}       # {uid: session_dict} — avoids re-reading Firestore on every action
_session_lock = threading.Lock()  # Protects _session_cache read-modify-write + Firestore writes

def _session_file_for_user(user_id=None):
    """Get session file path for a specific user"""
    uid = user_id or get_user_id()
    return os.path.join(DATA_DIR, f"session_{uid}.json")

def load_session(user_id=None):
    uid = user_id or get_user_id()
    with _session_lock:
        # In-memory cache hit — instant
        if uid in _session_cache:
            return _session_cache[uid]
    # Local file first (most recent writes go here), then Firestore
    path = _session_file_for_user(uid)
    if os.path.exists(path):
        try:
            with open(path) as f:
                content = f.read().strip()
                if content:
                    data = json.loads(content)
                    with _session_lock:
                        _session_cache[uid] = data
                    return data
        except (json.JSONDecodeError, IOError):
            pass
    # Firestore fallback (may have data from before last redeploy)
    fs_data = load_session_firestore(uid)
    if fs_data:
        with _session_lock:
            _session_cache[uid] = fs_data
        return fs_data
    default = {"accepted": [], "rejected": [], "current_index": 0, "previous_urls": []}
    with _session_lock:
        _session_cache[uid] = default
    return default

def save_session(session, user_id=None):
    uid = user_id or get_user_id()
    with _session_lock:
        _session_cache[uid] = session
    # Guardar en archivo local per-user (fast, synchronous — source of truth)
    try:
        path = _session_file_for_user(uid)
        tmp_path = path + ".tmp"
        with open(tmp_path, "w") as f:
            json.dump(session, f, ensure_ascii=False, indent=2)
            f.flush()
            os.fsync(f.fileno())
        os.replace(tmp_path, path)
    except Exception as e:
        print(f"⚠️ Error guardando session local: {e}")
    # Firestore write — deferred to background thread
    import copy
    session_copy = copy.deepcopy(session)  # Snapshot to avoid mutation during write
    def _write_firestore():
        try:
            save_session_firestore(session_copy, uid)
        except Exception as e:
            print(f"⚠️ Error guardando session en Firestore: {e}")
    threading.Thread(target=_write_firestore, daemon=True).start()

def invalidate_session_cache(user_id=None):
    """Clear session cache for a user (e.g., after reset)"""
    uid = user_id or get_user_id()
    with _session_lock:
        _session_cache.pop(uid, None)

# ─── Crawling ────────────────────────────────────────────────────────────

MAX_XML_SIZE = 5 * 1024 * 1024  # 5MB max for sitemap XML

# Countries that use dot as thousands separator, comma as decimal (or no decimal)
# CL, AR, CO, MX use: 29.990 or 29,990 (no decimals for CLP/COP/MXN/ARS)
# ES uses: 29,90 or 1.299,00 (comma = decimal)
_DOT_THOUSANDS_COUNTRIES = {"CL", "AR", "CO", "MX"}  # dot = thousands, no decimals
_COMMA_DECIMAL_COUNTRIES = {"ES"}  # comma = decimal, dot = thousands

def _normalize_price(price_str):
    """Normalize price string based on format detection.
    Chilean/LatAm: 29.990 → 29990 | Spanish: 29,90 → 29.90 | 1.299,00 → 1299.00
    """
    s = price_str.strip()
    # Detect format: if has comma after dot → comma is decimal (ES format: 1.299,00)
    if ',' in s and '.' in s:
        if s.rindex(',') > s.rindex('.'):
            # Comma is decimal: 1.299,00 → 1299.00
            return s.replace(".", "").replace(",", ".")
        else:
            # Dot is decimal: 1,299.00 → 1299.00
            return s.replace(",", "")
    elif ',' in s:
        # Only comma: could be "29,990" (thousands) or "29,90" (decimal)
        parts = s.split(',')
        if len(parts[-1]) == 2:
            # Two decimals after comma → comma is decimal: 29,90 → 29.90
            return s.replace(",", ".")
        else:
            # Comma is thousands: 29,990 → 29990
            return s.replace(",", "")
    elif '.' in s:
        # Only dot: could be "29.990" (thousands) or "29.90" (decimal)
        parts = s.split('.')
        if len(parts[-1]) == 3 and len(parts) > 1:
            # Three digits after last dot → dot is thousands: 29.990 → 29990
            return s.replace(".", "")
        elif len(parts[-1]) == 2:
            # Two digits after dot → dot is decimal: 29.90 → 29.90
            return s
        else:
            # Ambiguous — strip dots (assume thousands)
            return s.replace(".", "")
    return s

def safe_parse_xml(content):
    """Parse XML with size limit to prevent memory bombs."""

    if len(content) > MAX_XML_SIZE:
        print(f"  ⚠ XML too large ({len(content) / 1024 / 1024:.1f}MB), skipping")
        return None
    try:
        return ET.fromstring(content)
    except ET.ParseError:
        return None

def fetch_with_retry(url, max_retries=3, base_delay=2.0):
    """Fetch URL with exponential backoff retries"""
    for attempt in range(max_retries):
        try:
            resp = requests.get(url, headers=HEADERS, timeout=45, allow_redirects=True)
            if resp.status_code == 200:
                # Fix encoding for Latin American sites
                if resp.encoding and resp.encoding.lower() == 'iso-8859-1':
                    resp.encoding = resp.apparent_encoding
                return resp
            elif resp.status_code in (429, 500, 502, 503):
                # Rate limited or overloaded — wait longer
                wait = base_delay * (2 ** attempt) + 1
                print(f"    ⏳ {resp.status_code} — retrying in {wait:.0f}s (attempt {attempt + 1}/{max_retries})")
                time.sleep(wait)
            else:
                print(f"    ⚠ Status {resp.status_code}")
                return None
        except requests.exceptions.Timeout:
            wait = base_delay * (2 ** attempt)
            print(f"    ⏳ Timeout — retrying in {wait:.0f}s (attempt {attempt + 1}/{max_retries})")
            time.sleep(wait)
        except Exception as e:
            print(f"    ✗ Error: {e}")
            return None
    return None


def crawl_woocommerce(brand, progress_callback=None):
    """Fetch all products from a WooCommerce store via Store API + sitemap + HTML fallback"""
    products = []
    known_urls = set()
    base_url = brand["url"].rstrip("/")

    def log(msg):
        if progress_callback:
            progress_callback(msg)

    # ── PHASE 1: WooCommerce Store API ──────────────────────────────────
    page = 1
    while True:
        url = f"{base_url}/wp-json/wc/store/v1/products?per_page=100&page={page}"
        log(f"API página {page} de {brand['name']}... ({len(products)} productos)")

        resp = fetch_with_retry(url)
        if resp is None:
            break

        try:
            data = resp.json()
        except Exception:
            break

        if not data or not isinstance(data, list):
            break

        for p in data:
            images = p.get("images", [])
            image_url = images[0].get("src", "") if images else ""
            all_images = [img.get("src", "") for img in images[:5]]

            categories = p.get("categories", [])
            category = categories[0].get("name", "") if categories else ""

            prices = p.get("prices", {})
            price_raw = prices.get("price", "0")
            minor_unit = prices.get("currency_minor_unit", 0)
            try:
                price_int = int(price_raw)
                if minor_unit > 0:
                    price = str(price_int // (10 ** minor_unit))
                else:
                    price = str(price_int)
            except (ValueError, TypeError):
                price = price_raw

            tags = [t.get("name", "") for t in p.get("tags", [])]

            desc_html = p.get("short_description", "") or p.get("description", "") or ""
            description = ""
            if desc_html:
                soup = BeautifulSoup(desc_html, "html.parser")
                description = soup.get_text(separator=" ").strip()[:500]

            permalink = p.get("permalink", "")
            is_purchasable = p.get("is_purchasable", True)

            known_urls.add(permalink.rstrip("/").lower())
            products.append({
                "brand": brand["name"],
                "name": p.get("name", ""),
                "category": category if category else categorize(tags, p.get("name", "")),
                "price": price,
                "image_url": image_url,
                "all_images": all_images,
                "product_url": permalink,
                "description": description,
                "available": is_purchasable,
                "tags": tags[:8],
                "variants": [],
                "created_at": "",
            })

        if len(data) < 100:
            break
        page += 1
        time.sleep(2.0)

    log(f"{brand['name']}: {len(products)} vía API — buscando más en sitemap...")

    # ── PHASE 2: Product sitemap (catches products API might miss) ──────
    sitemap_urls = _fetch_woo_sitemap_urls(base_url)
    missing_urls = [u for u in sitemap_urls if u.rstrip("/").lower() not in known_urls]

    if missing_urls:
        log(f"{brand['name']}: {len(missing_urls)} productos extra en sitemap, scrapeando...")
        for i, purl in enumerate(missing_urls):
            log(f"{brand['name']}: scrapeando extra {i+1}/{len(missing_urls)}")
            scraped = _scrape_single_product_page(purl, brand)
            if scraped:
                known_urls.add(purl.rstrip("/").lower())
                products.append(scraped)
            time.sleep(1.5)

    # ── PHASE 3: HTML navigation (explore shop/catalog pages for any remaining) ──
    html_urls = _discover_product_urls_from_html(brand, known_urls)
    if html_urls:
        log(f"{brand['name']}: {len(html_urls)} productos extra en HTML, scrapeando...")
        for i, purl in enumerate(html_urls):
            log(f"{brand['name']}: scrapeando HTML {i+1}/{len(html_urls)}")
            scraped = _scrape_single_product_page(purl, brand)
            if scraped:
                known_urls.add(purl.rstrip("/").lower())
                products.append(scraped)
            time.sleep(1.5)

    log(f"✓ {brand['name']}: {len(products)} productos totales")
    return products


def _fetch_woo_sitemap_urls(base_url):
    """Try WooCommerce product sitemaps to discover all published product URLs"""


    sitemap_paths = [
        "/product-sitemap.xml",
        "/wp-sitemap-posts-product-1.xml",
    ]
    product_urls = []
    skip_suffixes = ("/tienda/", "/shop/", "/tienda", "/shop")

    for path in sitemap_paths:
        try:
            resp = requests.get(f"{base_url}{path}", headers=HEADERS, timeout=15)
            if resp.status_code != 200:
                continue
            root = safe_parse_xml(resp.content)
            if root is None:
                break
            ns = {'s': 'http://www.sitemaps.org/schemas/sitemap/0.9'}
            for url_el in root.findall('s:url', ns):
                loc = url_el.find('s:loc', ns)
                if loc is not None and loc.text:
                    u = loc.text.strip()
                    # Skip non-product pages (shop index, categories)
                    if not u.endswith(skip_suffixes) and u != base_url + "/":
                        product_urls.append(u)
            if product_urls:
                break  # Found a working sitemap
        except Exception as e:
            print(f"    Sitemap {path}: {e}")
            continue

    return product_urls


def _discover_product_urls_from_html(brand, already_known):
    """Navigate shop/catalog/collection pages (with pagination) to find product links not yet known"""
    base_url = brand["url"].rstrip("/")
    domain = brand.get("domain", "")

    listing_paths = [
        "/tienda", "/shop", "/productos", "/catalogo", "/collections/all",
        "/collections", "/showroom", "/product-category", "/categoria-producto",
        "/coleccion", "/coleccion-2025", "/coleccion-2026",
        "/nueva-coleccion", "/new-collection", "/all", "/"
    ]

    candidate_links = set()

    def _extract_product_links(soup):
        """Extract product links from a parsed page"""
        found = set()
        for a in soup.find_all("a", href=True):
            href = a["href"].split("?")[0].rstrip("/")
            if href.startswith("/"):
                href = base_url + href

            # Must be same domain
            if domain and domain not in href:
                continue
            if not href.startswith("http"):
                continue

            href_lower = href.lower()

            # Skip clearly non-product pages
            if any(skip in href_lower for skip in [
                "/cart", "/carrito", "/checkout", "/account", "/login", "/register",
                "/blog", "/pages/", "/policies", "/politica", "/terminos",
                "/nosotras", "/nosotros", "/about", "/contacto", "/contact",
                ".js", ".css", ".png", ".jpg", "#", "javascript:", "mailto:",
                "/wp-login", "/wp-admin", "/feed", "/reembolso", "/envio",
                "/collections/all", "/categories", "/search", "/mi-cuenta",
                "/page/", "/categoria-producto/", "/product-category/",
            ]):
                continue

            # Skip if it's the base URL itself
            if href.rstrip("/") == base_url:
                continue

            # Check if it has product-like patterns
            is_product_pattern = any(p in href_lower for p in [
                "/products/", "/producto/", "/product/", "/p/",
                "/item/", "/tienda/", "/shop/"
            ])

            # For WooCommerce: check if link has add-to-cart data attributes
            has_product_class = False
            if a.get("class"):
                classes = " ".join(a["class"]).lower()
                has_product_class = any(c in classes for c in [
                    "product", "woocommerce", "add_to_cart"
                ])

            # Accept product-pattern links, product-class links,
            # or root-level links from known WooCommerce sites
            if is_product_pattern or has_product_class:
                found.add(href)
            elif brand.get("platform") == "woocommerce":
                # WooCommerce flat slugs: accept if it's a simple /slug path
                path_part = href.replace(base_url, "").strip("/")
                if path_part and "/" not in path_part:
                    skip_slugs = [
                        "coleccion", "collection", "categoria", "category",
                        "tienda", "shop", "carrito", "cart", "checkout",
                        "nosotras", "nosotros", "about", "contacto", "contact",
                    ]
                    if not any(path_part.lower().startswith(s) for s in skip_slugs):
                        found.add(href)
        return found

    for path in listing_paths:
        try:
            # Try the listing page + up to 5 pagination pages
            for page_num in range(1, 6):
                if page_num == 1:
                    page_url = f"{base_url}{path}"
                else:
                    # WooCommerce pagination: /tienda/page/2/
                    page_url = f"{base_url}{path}/page/{page_num}/"

                resp = requests.get(page_url, headers=HEADERS, timeout=20, allow_redirects=True)
                if resp.status_code != 200:
                    break  # No more pagination pages

                soup = BeautifulSoup(resp.text, "html.parser")
                page_links = _extract_product_links(soup)
                candidate_links.update(page_links)

                # Check if there's a next page
                has_next = bool(
                    soup.find("a", class_="next") or
                    soup.find("a", attrs={"rel": "next"}) or
                    soup.find("link", attrs={"rel": "next"}) or
                    soup.select_one(".woocommerce-pagination .next, .pagination .next, a.next.page-numbers")
                )
                if not has_next:
                    break
                time.sleep(1.0)

            time.sleep(0.5)
        except Exception:
            continue

    # Also check category pages if we found category links
    category_urls = set()
    for link in list(candidate_links):
        link_lower = link.lower()
        if any(cat in link_lower for cat in ["/categoria-producto/", "/product-category/", "/product_cat/"]):
            category_urls.add(link)
            candidate_links.discard(link)  # It's a category, not a product

    for cat_url in category_urls:
        try:
            resp = requests.get(cat_url, headers=HEADERS, timeout=20, allow_redirects=True)
            if resp.status_code == 200:
                soup = BeautifulSoup(resp.text, "html.parser")
                candidate_links.update(_extract_product_links(soup))
            time.sleep(1.0)
        except Exception:
            continue

    # Remove already known
    return [u for u in candidate_links if u.rstrip("/").lower() not in already_known]


def _scrape_single_product_page(url, brand):
    """Scrape a single product page for details"""
    base_url = brand["url"].rstrip("/")
    try:
        resp = requests.get(url, headers=HEADERS, timeout=20, allow_redirects=True)
        if resp.status_code != 200:
            return None

        soup = BeautifulSoup(resp.text, "html.parser")

        # Verify it's actually a product page (multiple signals)
        has_price = bool(soup.find(class_="price") or soup.find(attrs={"itemprop": "price"})
                         or soup.select_one("[class*='price']"))
        has_cart = bool(soup.find(class_="single_add_to_cart_button")
                        or soup.find("button", {"name": "add-to-cart"})
                        or soup.find("form", {"action": "/cart/add"})  # Shopify
                        or soup.find("form", class_="product-form"))
        og_type = soup.find("meta", property="og:type")
        og_type_val = og_type.get("content", "").lower() if og_type else ""
        is_og_product = og_type_val in ("product", "og:product")
        is_og_website = og_type_val == "website"  # Definitely a listing/collection page, not a product
        has_product_schema = bool(soup.find(attrs={"itemtype": lambda v: v and "Product" in v}))

        # A page with og:type=website is a collection/listing page — never a product page.
        # A page with only has_price (no cart, no og:product, no schema) is likely a listing page.
        if is_og_website and not has_product_schema:
            return None
        if not has_price and not has_cart and not is_og_product and not has_product_schema:
            return None
        # has_price alone on a page without a cart or product schema is unreliable
        # (listing pages show prices). Require at least one stronger signal.
        if has_price and not has_cart and not is_og_product and not has_product_schema:
            # Accept only if there's a single price element (not multiple → listing page)
            price_els = soup.select("[class*='price']")
            if len(price_els) > 3:
                return None  # Too many price elements = listing/category page

        # Extract name
        name = ""
        for selector in ["h1.product_title", "h1", ".product-title", "[itemprop='name']"]:
            el = soup.select_one(selector)
            if el:
                name = el.get_text(strip=True)
                break
        if not name:
            og_title = soup.find("meta", property="og:title")
            name = og_title["content"].split(" - ")[0].strip() if og_title else url.split("/")[-1].replace("-", " ").title()

        # Extract image
        image_url = ""
        og_img = soup.find("meta", property="og:image")
        if og_img and og_img.get("content"):
            image_url = og_img["content"]
        else:
            for sel in [".woocommerce-product-gallery img", ".product-image img", "img.wp-post-image"]:
                img = soup.select_one(sel)
                if img and img.get("src"):
                    src = img["src"]
                    if src.startswith("//"):
                        src = "https:" + src
                    elif src.startswith("/"):
                        src = base_url + src
                    image_url = src
                    break

        # Extract price (works across WooCommerce, Shopify, generic)

        price = ""
        for price_sel in [
            # Meta tags first — most reliable, no risk of matching listing prices
            "meta[property='product:price:amount']",
            "meta[name='product_price']",
            # Structured data attributes
            "[itemprop='price']",
            # Specific platform price elements (before generic [class*='price'])
            ".price .woocommerce-Price-amount",
            ".price ins .woocommerce-Price-amount",  # WooCommerce sale price
            ".price", ".product-price", ".current-price",
            # Generic fallback — last resort (can match listing pages)
            "[class*='price']",
        ]:
            price_el = soup.select_one(price_sel)
            if price_el:
                if price_el.name == "meta":
                    raw = price_el.get("content", "")
                    # Strip decimal part for LatAm (e.g., "149990.0" → "149990")
                    try:
                        price = str(int(float(raw))) if raw else ""
                    except (ValueError, TypeError):
                        price = raw
                else:
                    price_text = price_el.get_text(strip=True)
                    nums = re.findall(r'[\d.,]+', price_text)
                    if nums:
                        price = _normalize_price(nums[0])
                # Skip "0" — likely a placeholder/missing price, try next selector
                if price and price != "0":
                    break

        # Extract description
        description = ""
        og_desc = soup.find("meta", property="og:description")
        if og_desc and og_desc.get("content"):
            description = og_desc["content"][:500]

        # All images (WooCommerce, Shopify, generic)
        all_images = [image_url] if image_url else []
        for img in soup.select(".woocommerce-product-gallery img, .product-images img, .thumbnails img, .product-gallery img, .product__media img, [class*='product'] img"):
            src = img.get("data-large_image") or img.get("data-src") or img.get("src", "")
            if src.startswith("//"):
                src = "https:" + src
            elif src.startswith("/"):
                src = base_url + src
            if src and src not in all_images:
                all_images.append(src)
            if len(all_images) >= 5:
                break

        return {
            "brand": brand["name"],
            "name": name,
            "category": categorize([], name),
            "price": price,
            "image_url": image_url,
            "all_images": all_images,
            "product_url": url,
            "description": description,
            "available": True,
            "tags": [],
            "variants": [],
            "created_at": "",
        }
    except Exception as e:
        print(f"    Error scraping {url}: {e}")
        return None


def detect_platform(brand, progress_callback=None):
    """Auto-detect store platform: Shopify, WooCommerce, Jumpseller, Tiendanube, VTEX, PrestaShop, or HTML fallback"""
    if brand.get("platform"):
        return brand["platform"]

    base_url = brand["url"].rstrip("/")

    def log(msg):
        if progress_callback:
            progress_callback(msg)

    # 1. Try Shopify API
    log(f"Probando Shopify en {brand['name']}...")
    try:
        resp = requests.get(f"{base_url}/products.json?limit=1", headers=HEADERS, timeout=15, allow_redirects=True)
        if resp.status_code == 200:
            data = resp.json()
            if "products" in data:
                return "shopify"
    except Exception:
        pass

    # 2. Try WooCommerce Store API
    log(f"Probando WooCommerce en {brand['name']}...")
    try:
        resp = requests.get(f"{base_url}/wp-json/wc/store/v1/products?per_page=1", headers=HEADERS, timeout=15, allow_redirects=True)
        if resp.status_code == 200:
            data = resp.json()
            if isinstance(data, list) and len(data) > 0:
                return "woocommerce"
    except Exception:
        pass

    # 3-5. Detect from homepage HTML (single request)
    log(f"Analizando HTML de {brand['name']}...")
    try:
        resp = requests.get(base_url, headers=HEADERS, timeout=15, allow_redirects=True)
        if resp.status_code == 200:
            html = resp.text.lower()
            if "jumpseller" in html:
                return "jumpseller"
            if "tiendanube" in html or "nuvemshop" in html:
                return "tiendanube"
            if "vtex" in html or "vteximg" in html:
                return "vtex"
            if "prestashop" in html:
                return "prestashop"
            # Magento: specific markers (not just "mage" which matches "image")
            if "magento" in html or "mage-init" in html or "mage/cookies" in html:
                return "magento"
            if '/product' in html or '/productos' in html or '/collections' in html:
                return "html_scrape"
    except Exception:
        pass

    # 6. Last resort — try HTML scraping anyway
    return "html_scrape"


def crawl_html_scrape(brand, progress_callback=None):
    """Scrape products from HTML pages — uses sitemap + HTML navigation + page scraping"""
    base_url = brand["url"].rstrip("/")
    known_urls = set()
    products = []

    def log(msg):
        if progress_callback:
            progress_callback(msg)

    log(f"Scraping HTML de {brand['name']}...")

    # ── PHASE 1: Try sitemaps first (most reliable for any platform) ────
    sitemap_urls = _fetch_generic_sitemap_urls(base_url, brand.get("domain", ""))
    if sitemap_urls:
        # Use a shorter delay for larger sitemaps to stay within the 3-minute timeout
        sitemap_delay = 0.8 if len(sitemap_urls) > 60 else 1.5
        log(f"{brand['name']}: {len(sitemap_urls)} URLs en sitemap, scrapeando...")
        for i, purl in enumerate(sitemap_urls):
            if i % 5 == 0:
                log(f"{brand['name']}: producto {i+1}/{len(sitemap_urls)} (sitemap)")
            scraped = _scrape_single_product_page(purl, brand)
            if scraped:
                known_urls.add(purl.rstrip("/").lower())
                products.append(scraped)
            time.sleep(sitemap_delay)

    # ── PHASE 2: HTML navigation (explore shop/catalog pages) ───────────
    log(f"{brand['name']}: buscando más en páginas de catálogo...")

    listing_paths = [
        "/collections/all", "/products", "/productos", "/tienda",
        "/shop", "/catalogo", "/collection/all", "/collections",
        "/categoria-producto", "/product-category", "/showroom",
        "/coleccion", "/coleccion-2025", "/coleccion-2026",
        "/new-arrivals", "/novedades", "/all", "/"
    ]

    product_links = set()

    for path in listing_paths:
        try:
            url = f"{base_url}{path}"
            resp = requests.get(url, headers=HEADERS, timeout=20, allow_redirects=True)
            if resp.status_code != 200:
                continue

            soup = BeautifulSoup(resp.text, "html.parser")

            for a in soup.find_all("a", href=True):
                href = a["href"]
                if href.startswith("/"):
                    href = base_url + href
                elif not href.startswith("http"):
                    continue

                href_lower = href.lower()
                if any(skip in href_lower for skip in [
                    "/cart", "/carrito", "/checkout", "/account", "/login", "/register",
                    "/blog", "/pages/", "/policies", "/politica", "/terminos",
                    "/nosotras", "/nosotros", "/about", "/contacto", "/contact",
                    ".js", ".css", ".png", ".jpg", "#", "javascript:", "mailto:",
                    "/wp-login", "/wp-admin", "/feed", "/reembolso", "/envio",
                    "/collections/all", "/categories", "/search", "/mi-cuenta",
                ]):
                    continue

                if brand["domain"] not in href:
                    continue

                href_clean = href.split("?")[0].rstrip("/")
                if href_clean.rstrip("/").lower() in known_urls:
                    continue

                if any(pattern in href_lower for pattern in [
                    "/products/", "/producto/", "/product/", "/p/",
                    "/item/", "/tienda/", "/shop/"
                ]):
                    product_links.add(href_clean)

            if product_links:
                break

            time.sleep(1.5)
        except Exception as e:
            print(f"    Error scraping {path}: {e}")
            continue

    # Paginated listing
    page = 2
    while len(product_links) > 0 and page <= 10:
        try:
            for pag_url in [
                f"{base_url}/collections/all?page={page}",
                f"{base_url}/products?page={page}",
                f"{base_url}/tienda/page/{page}/",
                f"{base_url}/shop/page/{page}/",
            ]:
                resp = requests.get(pag_url, headers=HEADERS, timeout=15, allow_redirects=True)
                if resp.status_code != 200:
                    continue
                soup = BeautifulSoup(resp.text, "html.parser")
                new_links = 0
                for a in soup.find_all("a", href=True):
                    href = a["href"]
                    if href.startswith("/"):
                        href = base_url + href
                    href_clean = href.split("?")[0].rstrip("/")
                    if brand["domain"] in href and any(p in href.lower() for p in ["/products/", "/producto/", "/product/"]):
                        if href_clean not in product_links and href_clean.rstrip("/").lower() not in known_urls:
                            product_links.add(href_clean)
                            new_links += 1
                if new_links > 0:
                    break
                else:
                    break
            page += 1
            time.sleep(2.0)
        except Exception:
            break

    # Scrape product links found via HTML
    if product_links:
        log(f"{brand['name']}: {len(product_links)} links encontrados en HTML, extrayendo datos...")
        for i, link in enumerate(sorted(product_links)):
            if i % 5 == 0:
                log(f"{brand['name']}: producto {i+1}/{len(product_links)} (HTML)")
            scraped = _scrape_single_product_page(link, brand)
            if scraped:
                known_urls.add(link.rstrip("/").lower())
                products.append(scraped)
            time.sleep(1.5)

    if not products:
        print(f"    No products found for {brand['name']}")

    log(f"✓ {brand['name']}: {len(products)} productos totales")
    return products


def _fetch_generic_sitemap_urls(base_url, domain=""):
    """Try common sitemap locations to discover product URLs for any platform"""


    sitemap_paths = [
        "/product-sitemap.xml",
        "/wp-sitemap-posts-product-1.xml",
        "/sitemap_products_1.xml",
        "/sitemap.xml",
    ]
    product_urls = []
    ns = {'s': 'http://www.sitemaps.org/schemas/sitemap/0.9'}

    for path in sitemap_paths:
        try:
            resp = requests.get(f"{base_url}{path}", headers=HEADERS, timeout=15)
            if resp.status_code != 200:
                continue

            root = safe_parse_xml(resp.content)
            if root is None:
                break

            # Check if it's a sitemap index (contains <sitemap> elements)
            sub_sitemaps = root.findall('s:sitemap', ns)
            if sub_sitemaps:
                # Find product-related sub-sitemaps
                for sm in sub_sitemaps:
                    loc = sm.find('s:loc', ns)
                    if loc is not None and 'product' in loc.text.lower():
                        try:
                            sm_resp = requests.get(loc.text, headers=HEADERS, timeout=15)
                            if sm_resp.status_code != 200:
                                continue
                            sm_root = safe_parse_xml(sm_resp.content)
                            if sm_root is None:
                                continue
                            for url_el in sm_root.findall('s:url', ns):
                                loc2 = url_el.find('s:loc', ns)
                                if loc2 is not None:
                                    u = loc2.text.strip()
                                    if '/product' in u.lower() or (domain and domain in u):
                                        product_urls.append(u)
                        except Exception:
                            continue
                if product_urls:
                    break
            else:
                # Direct sitemap with <url> elements
                for url_el in root.findall('s:url', ns):
                    loc = url_el.find('s:loc', ns)
                    if loc is not None:
                        u = loc.text.strip()
                        # For product-specific sitemaps, include all URLs except shop index
                        if 'product' in path:
                            skip_suffixes = ("/tienda/", "/shop/", "/tienda", "/shop")
                            if not u.endswith(skip_suffixes) and u.rstrip("/") != base_url:
                                product_urls.append(u)
                        elif '/product' in u.lower():
                            product_urls.append(u)

            if product_urls:
                break
        except Exception:
            continue

    return product_urls


def crawl_jumpseller(brand, progress_callback=None):
    """Crawl a Jumpseller store — uses sitemap + homepage product detection.

    Jumpseller stores often use flat URLs (e.g., /miuccia, /eloise) without
    /products/ prefix, so the generic HTML scraper misses them. This crawler
    parses the sitemap and tries each non-page URL as a potential product.
    """
    base_url = brand["url"].rstrip("/")
    domain = brand.get("domain", "")
    products = []
    known_urls = set()

    def log(msg):
        if progress_callback:
            progress_callback(msg)

    log(f"Crawleando Jumpseller: {brand['name']}...")

    # Non-product path segments to skip
    SKIP_SLUGS = {
        "blog", "contact", "contacto", "conocenos", "about", "nosotras", "nosotros",
        "cart", "carrito", "checkout", "customer", "login", "register",
        "despachos", "envio", "envios", "politica-de-privacidad", "politica-de-reembolso",
        "politicas-de-cambio", "terminos-y-condiciones", "terminos", "condiciones",
        "como-cuidar-mi-zapato", "faq", "preguntas-frecuentes",
        "pages", "policies", "search", "collections", "categories",
        "informacion", "information", "venta-especial", "rebajas", "sale",
    }

    # ── PHASE 1: Sitemap — try to scrape each non-page URL ──────────

    candidate_urls = []

    try:
        resp = requests.get(f"{base_url}/sitemap.xml", headers=HEADERS, timeout=15)
        if resp.status_code == 200:
            root = safe_parse_xml(resp.content)
            if root is not None:
                ns = {'s': 'http://www.sitemaps.org/schemas/sitemap/0.9'}
                for url_el in root.findall('s:url', ns):
                    loc = url_el.find('s:loc', ns)
                    if loc is None:
                        continue
                    u = loc.text.strip().rstrip("/")
                    if u == base_url:
                        continue
                    # Only root-level URLs (single path segment) — sub-pages like
                    # /37-1/disponibilidad-por-talla-37 are size/availability pages, not products
                    path_parts = [s for s in u.replace(base_url, "").strip("/").split("/") if s]
                    if len(path_parts) != 1:
                        continue
                    slug = path_parts[0].lower()
                    if slug and slug not in SKIP_SLUGS:
                        candidate_urls.append(u)
    except Exception as e:
        print(f"    Jumpseller sitemap error: {e}")

    if candidate_urls:
        # Cap at 200 candidates to stay within the 3-minute per-brand timeout
        if len(candidate_urls) > 200:
            print(f"    {brand['name']}: capping {len(candidate_urls)} candidates to 200")
            candidate_urls = candidate_urls[:200]
        log(f"{brand['name']}: {len(candidate_urls)} URLs en sitemap, verificando cuáles son productos...")
        for i, purl in enumerate(candidate_urls):
            if i % 5 == 0:
                log(f"{brand['name']}: verificando {i+1}/{len(candidate_urls)}")
            scraped = _scrape_single_product_page(purl, brand)
            if scraped:
                known_urls.add(purl.rstrip("/").lower())
                products.append(scraped)
            time.sleep(0.5)  # Reduced from 1.0s to stay within timeout budget

    # ── PHASE 2: Homepage links — catch anything not in sitemap ──────
    if not products:
        log(f"{brand['name']}: buscando productos en homepage...")
        try:
            resp = requests.get(base_url, headers=HEADERS, timeout=20)
            if resp.status_code == 200:
                soup = BeautifulSoup(resp.text, "html.parser")
                for a in soup.find_all("a", href=True):
                    href = a["href"].split("?")[0].rstrip("/")
                    if href.startswith("/"):
                        href = base_url + href
                    if not href.startswith("http") or (domain and domain not in href):
                        continue
                    slug = href.replace(base_url, "").strip("/").split("/")[0].lower()
                    if slug and slug not in SKIP_SLUGS and href.rstrip("/").lower() not in known_urls:
                        candidate_urls.append(href)

                for purl in candidate_urls:
                    if purl.rstrip("/").lower() in known_urls:
                        continue
                    scraped = _scrape_single_product_page(purl, brand)
                    if scraped:
                        known_urls.add(purl.rstrip("/").lower())
                        products.append(scraped)
                    time.sleep(1.0)
        except Exception as e:
            print(f"    Homepage scrape error: {e}")

    log(f"✓ {brand['name']}: {len(products)} productos totales")
    return products


def crawl_tiendanube(brand, progress_callback=None):
    """Crawl a Tiendanube store via HTML"""
    return crawl_html_scrape(brand, progress_callback)


def crawl_brand(brand, progress_callback=None):
    """Fetch all products — auto-detects platform and uses the right method"""
    # Check robots.txt first
    if not is_crawl_allowed(brand["url"]):
        if progress_callback:
            progress_callback(f"⚠ {brand['name']}: bloqueado por robots.txt")
        print(f"  🚫 {brand['name']}: blocked by robots.txt")
        return []

    # Normalize URL: follow redirects to get canonical base URL (e.g., non-www → www or vice versa).
    # Also handles 403 Cloudflare blocks by trying the alternate www variant.
    try:
        test_resp = requests.get(brand["url"], headers=HEADERS, timeout=10, allow_redirects=True)
        if test_resp.status_code == 403:
            # Try alternate www variant
            if "://www." in brand["url"]:
                alt_url = brand["url"].replace("://www.", "://", 1)
            else:
                alt_url = brand["url"].replace("://", "://www.", 1)
            try:
                alt_resp = requests.get(alt_url, headers=HEADERS, timeout=10, allow_redirects=True)
                if alt_resp.status_code == 200:
                    print(f"  🔄 {brand['name']}: swapping URL {brand['url']} → {alt_url} (403 on original)")
                    brand = dict(brand)
                    brand["url"] = alt_url.rstrip("/")
            except Exception:
                pass
        elif test_resp.status_code == 200:
            # Follow redirect to canonical URL (e.g., wearearde.cl → www.wearearde.cl)
            canonical = test_resp.url.rstrip("/")
            if canonical != brand["url"].rstrip("/") and brand.get("domain", "") in canonical:
                print(f"  🔄 {brand['name']}: canonical URL {brand['url']} → {canonical}")
                brand = dict(brand)
                brand["url"] = canonical
    except Exception:
        pass

    if progress_callback:
        progress_callback(f"Detectando plataforma de {brand['name']}...")

    platform = detect_platform(brand, progress_callback)

    if progress_callback:
        progress_callback(f"{brand['name']} → {platform.upper()} detectado")

    crawlers = {
        "shopify": crawl_shopify,
        "woocommerce": crawl_woocommerce,
        "jumpseller": crawl_jumpseller,
        "tiendanube": crawl_tiendanube,
        "vtex": crawl_html_scrape,
        "prestashop": crawl_html_scrape,
        "magento": crawl_html_scrape,
        "html_scrape": crawl_html_scrape,
    }

    crawler = crawlers.get(platform, crawl_html_scrape)
    products = crawler(brand, progress_callback)

    # Detect JS-only sites: got 0 products but page loaded OK
    if not products and platform == "html_scrape":
        try:
            resp = requests.get(brand["url"], headers=HEADERS, timeout=15)
            body = resp.text.lower()
            js_signals = ["__next", "react-root", "nuxt", "vue-app", "ng-app",
                          "window.__initial", "hydrate", "data-reactroot"]
            if any(sig in body for sig in js_signals):
                msg = f"⚠ {brand['name']}: sitio JavaScript (SPA) — no se puede scrapear sin navegador"
                if progress_callback:
                    progress_callback(msg)
                print(f"  🔧 {brand['name']}: detected JS-only site (SPA)")
                log_error_to_firestore("js_only_site", msg, {"domain": brand.get("domain", "")})
        except Exception:
            pass

    return products


def crawl_shopify(brand, progress_callback=None):
    """Fetch all products from a Shopify store via API + sitemap + HTML fallback"""
    products = []
    known_urls = set()
    base_url = brand["url"].rstrip("/")

    def log(msg):
        if progress_callback:
            progress_callback(msg)

    # ── PHASE 1: Shopify JSON API ───────────────────────────────────────
    page = 1
    MAX_SHOPIFY_PAGES = 50  # Cap at 12,500 products
    while page <= MAX_SHOPIFY_PAGES:
        url = f"{base_url}/products.json?limit=250&page={page}"
        log(f"API página {page} de {brand['name']}... ({len(products)} productos)")

        resp = fetch_with_retry(url)
        if resp is None:
            break

        try:
            data = resp.json()
        except Exception:
            print(f"    ✗ Invalid JSON from {brand['name']} page {page}")
            break

        page_products = data.get("products", [])
        if not page_products:
            break

        for p in page_products:
            images = p.get("images", [])
            image_url = images[0]["src"] if images else ""
            all_images = [img["src"] for img in images[:5]]

            product_type = p.get("product_type", "").strip()
            tags = p.get("tags", [])
            if isinstance(tags, str):
                tags = [t.strip() for t in tags.split(",")]

            body_html = p.get("body_html", "") or ""
            description = ""
            if body_html:
                soup = BeautifulSoup(body_html, "html.parser")
                description = soup.get_text(separator=" ").strip()[:500]

            variants = p.get("variants", [])
            available_variants = [v for v in variants if v.get("available", False)]
            price_variant = available_variants[0] if available_variants else (variants[0] if variants else {})
            price = price_variant.get("price", "")
            available = len(available_variants) > 0

            handle = p.get("handle", "")
            product_url = f"{base_url}/products/{handle}" if handle else ""

            variant_titles = []
            seen_titles = set()
            for v in variants:
                t = v.get("title", "").strip()
                if t and t != "Default Title" and t not in seen_titles:
                    seen_titles.add(t)
                    variant_titles.append(t)
            variant_titles = variant_titles[:6]

            known_urls.add(product_url.rstrip("/").lower())
            products.append({
                "brand": brand["name"],
                "name": p.get("title", ""),
                "category": product_type if product_type else categorize(tags, p.get("title", "")),
                "price": price,
                "image_url": image_url,
                "all_images": all_images,
                "product_url": product_url,
                "description": description,
                "available": available,
                "tags": tags[:8],
                "variants": variant_titles[:6],
                "created_at": p.get("created_at", ""),
            })

        page += 1
        time.sleep(2.0)

    log(f"{brand['name']}: {len(products)} vía API — buscando más en sitemap/HTML...")

    # ── PHASE 2: Shopify sitemap ────────────────────────────────────────
    sitemap_urls = _fetch_shopify_sitemap_urls(base_url)
    missing_urls = [u for u in sitemap_urls if u.rstrip("/").lower() not in known_urls]

    if missing_urls:
        log(f"{brand['name']}: {len(missing_urls)} productos extra en sitemap, scrapeando...")
        for i, purl in enumerate(missing_urls):
            log(f"{brand['name']}: scrapeando extra {i+1}/{len(missing_urls)}")
            scraped = _scrape_single_product_page(purl, brand)
            if scraped:
                known_urls.add(purl.rstrip("/").lower())
                products.append(scraped)
            time.sleep(1.5)

    # ── PHASE 3: HTML navigation ────────────────────────────────────────
    html_urls = _discover_product_urls_from_html(brand, known_urls)
    if html_urls:
        log(f"{brand['name']}: {len(html_urls)} productos extra en HTML, scrapeando...")
        for i, purl in enumerate(html_urls):
            log(f"{brand['name']}: scrapeando HTML {i+1}/{len(html_urls)}")
            scraped = _scrape_single_product_page(purl, brand)
            if scraped:
                known_urls.add(purl.rstrip("/").lower())
                products.append(scraped)
            time.sleep(1.5)

    log(f"✓ {brand['name']}: {len(products)} productos totales")
    return products


def _fetch_shopify_sitemap_urls(base_url):
    """Fetch product URLs from Shopify sitemap"""


    product_urls = []
    try:
        # Shopify main sitemap index
        resp = requests.get(f"{base_url}/sitemap.xml", headers=HEADERS, timeout=15)
        if resp.status_code != 200:
            return []

        root = safe_parse_xml(resp.content)
        if root is None:
            return []
        ns = {'s': 'http://www.sitemaps.org/schemas/sitemap/0.9'}

        # Find product sitemap URLs in the index
        product_sitemap_urls = []
        for sitemap in root.findall('s:sitemap', ns):
            loc = sitemap.find('s:loc', ns)
            if loc is not None and 'products' in loc.text.lower():
                product_sitemap_urls.append(loc.text)

        # If no index (direct sitemap), check for product URLs directly
        if not product_sitemap_urls:
            for url_el in root.findall('s:url', ns):
                loc = url_el.find('s:loc', ns)
                if loc is not None and '/products/' in loc.text:
                    product_urls.append(loc.text.strip())
            return product_urls

        # Fetch each product sitemap
        for sm_url in product_sitemap_urls:
            try:
                resp = requests.get(sm_url, headers=HEADERS, timeout=15)
                if resp.status_code != 200:
                    continue
                sm_root = safe_parse_xml(resp.content)
                if sm_root is None:
                    continue
                for url_el in sm_root.findall('s:url', ns):
                    loc = url_el.find('s:loc', ns)
                    if loc is not None and '/products/' in loc.text:
                        product_urls.append(loc.text.strip())
            except Exception:
                continue

    except Exception as e:
        print(f"    Shopify sitemap error: {e}")

    return product_urls


def categorize(tags, title):
    text = " ".join(tags).lower() + " " + title.lower()
    cats = {
        "Vestido": ["vestido", "dress"],
        "Blazer": ["blazer", "chaqueta", "jacket"],
        "Pantalón": ["pantalón", "pantalon", "pants"],
        "Falda": ["falda", "skirt", "pollera"],
        "Blusa": ["blusa", "top", "camiseta", "camisa", "shirt"],
        "Zapatos": ["zapato", "bota", "botin", "sandalia", "shoe", "boot", "mocasín",
                    "mocasin", "ballerina", "bailarina", "sapatilha", "taco", "plataforma",
                    "stiletto", "loafer", "oxford", "sneaker", "zapatilla"],
        "Bolso": ["bolso", "cartera", "bag", "tote", "clutch", "riñonera", "mochila"],
        "Accesorio": ["accesorio", "collar", "arete", "cinturón", "pañuelo",
                      "arnes", "arnés", "hebilla", "pulsera", "anillo", "aros", "joyas"],
        "Shorts": ["short", "shorts"],
        "Jeans": ["jean", "jeans", "denim"],
        "Pijama": ["pijama", "pajama"],
    }
    for cat, kws in cats.items():
        for kw in kws:
            if kw in text:
                return cat
    return "General"


def filter_unwanted_products(products):
    """Remove kids products and non-fashion items"""


    EXCLUDE_KEYWORDS = [
        # Kids
        "niño", "niña", "niños", "niñas", "kids", "kid", "child", "children",
        "bebé", "bebe", "baby", "infantil", "junior",
        # Non-fashion items
        "bolsa de compras", "bolsa regalo", "gift bag", "shopping bag",
        "gift card", "tarjeta de regalo", "giftcard", "tarjeta regalo",
        "embalaje", "packaging", "envoltorio", "wrapping",
        "vela", "candle", "incienso", "incense", "difusor",
        "sticker", "llavero", "keychain", "imán", "magnet",
        "taza", "mug", "plato", "plate",
        "libro", "revista", "magazine",
        "mascota", "perro", "gato",
    ]

    # Build regex pattern with word boundaries to avoid false positives
    # e.g. "pet" should NOT match "petite", "book" should NOT match "facebook"
    pattern = re.compile(
        r'\b(' + '|'.join(re.escape(kw) for kw in EXCLUDE_KEYWORDS) + r')\b',
        re.IGNORECASE
    )

    filtered = []
    for p in products:
        # Check name, tags, and category
        text = (p.get("name", "") + " " + p.get("category", "") + " " +
                " ".join(p.get("tags", []))).lower()

        if pattern.search(text):
            continue

        filtered.append(p)

    removed = len(products) - len(filtered)
    if removed > 0:
        print(f"  Filtered {removed} unwanted products (kids/non-fashion)")
    return filtered


def deduplicate_products(products):
    """
    Remove duplicate products based on:
    1. Same product URL (canonical dedup key)
    2. Same base name with different size/color suffix (variant dedup)
    Returns deduplicated list.
    """
    seen_urls = set()
    seen_names = set()
    unique = []



    for p in products:
        # Primary dedup: by product URL (most reliable)
        url = p.get("product_url", "").split("?")[0].rstrip("/").lower()
        if url and url in seen_urls:
            continue

        # Secondary dedup: normalize name to catch size/color variants
        raw_name = p.get("name", "")
        clean_name = re.sub(r'\s*/\s*(XS|S|M|L|XL|XXL|XXXL|\d{2,3})\s*$', '', raw_name, flags=re.IGNORECASE)
        clean_name = re.sub(r'\s*-\s*(Negro|Blanco|Rojo|Azul|Verde|Beige|Crudo|Café|Gris|Rosa|Nude|Burdeo|Camel|Mostaza|Terracota|Ivory|Black|White|Red|Blue|Green)\s*$', '', clean_name, flags=re.IGNORECASE)
        clean_name = re.sub(r'\s*talla\s*\d+\s*$', '', clean_name, flags=re.IGNORECASE)
        name_key = f"{p['brand']}|{clean_name.strip().lower()}"

        if name_key in seen_names:
            continue

        if url:
            seen_urls.add(url)
        seen_names.add(name_key)
        unique.append(p)

    return unique


# Global progress state — simple dict, read by polling endpoint
_default_progress = {"status": "idle", "message": "", "brand_idx": 0, "brand_total": 0, "user": "",
                     "products_found": 0, "current_brand": "", "done": False, "failed_brands": []}
# Single global progress dict — simpler and guaranteed to work
# Multi-user crawling is already prevented by crawl_lock
crawl_progress = dict(_default_progress)

def get_crawl_progress(uid=None):
    """Return progress. If another user is crawling, show a waiting message."""
    progress_user = crawl_progress.get("user", "")
    if uid and progress_user and progress_user != uid and not crawl_progress.get("done", True):
        return {"status": "waiting", "message": f"Otro usuario está crawleando ({progress_user})",
                "done": False, "brand_idx": 0, "brand_total": 0, "products_found": 0}
    return crawl_progress
crawl_lock = threading.Lock()  # Prevent concurrent crawls
crawl_lock_time = 0  # Timestamp when lock was acquired (auto-expire after 10 min)
crawl_cancel_event = threading.Event()  # Signal crawl thread to stop

# In-memory buffer for the last generated spreadsheet (survives ephemeral filesystem)
generated_xlsx_per_user = {}  # {user_id: BytesIO buffer} — max 10 entries
_XLSX_CACHE_MAX = 10
_last_cleanup_time = 0  # Throttle health check cleanup to once per hour
_global_cache_lock = threading.Lock()  # Protects _products_cache and generated_xlsx_per_user
_products_cache = {}  # {cache_file_path: (mtime, products_list)} — max 5 entries
_PRODUCTS_CACHE_MAX = 5

def _get_cached_products(country):
    """Load products from in-memory cache or file. Avoids re-parsing JSON on every /curate/next call."""
    cache_path = get_cache_file_for_country(country)
    if os.path.exists(cache_path):
        mtime = os.path.getmtime(cache_path)
        with _global_cache_lock:
            if cache_path in _products_cache and _products_cache[cache_path][0] == mtime:
                return _products_cache[cache_path][1]
        try:
            with open(cache_path) as f:
                products = json.load(f).get("products", [])
            with _global_cache_lock:
                if len(_products_cache) >= _PRODUCTS_CACHE_MAX:
                    oldest_key = next(iter(_products_cache))
                    del _products_cache[oldest_key]
                _products_cache[cache_path] = (mtime, products)
            return products
        except Exception:
            pass
    # Fallback to Firestore
    fs_products = load_cache_firestore(country)
    if fs_products:
        return fs_products
    return None


def crawl_all(brands=None, cache_file=None, progress=None, country=None, user_id=None):
    """Crawl all brands, deduplicate, and cache results.

    Uses a global dict for progress (read by polling endpoint).
    Saves partial cache after each brand for resilience.
    """
    global crawl_progress
    if progress is None:
        progress = crawl_progress
    # Ensure we're mutating the global dict that the polling endpoint reads
    crawl_progress = progress
    if brands is None:
        brands = load_active_brands()
    if cache_file is None:
        cache_file = CRAWL_CACHE

    crawl_cancel_event.clear()  # Reset cancel signal from previous cancellation
    # Replace contents atomically to avoid readers seeing empty dict
    new_state = {
        "status": "running", "message": "Iniciando...",
        "brand_idx": 0, "brand_total": len(brands),
        "products_found": 0, "current_brand": "", "done": False,
        "failed_brands": [], "user": user_id or "unknown"
    }
    for k in list(crawl_progress.keys()):
        if k not in new_state:
            del crawl_progress[k]
    crawl_progress.update(new_state)

    # Load previous cache to preserve data for brands that fail
    prev_products_by_brand = {}
    if os.path.exists(cache_file):
        try:
            with open(cache_file) as f:
                prev = json.load(f).get("products", [])
                for p in prev:
                    b = p.get("brand", "")
                    if b not in prev_products_by_brand:
                        prev_products_by_brand[b] = []
                    prev_products_by_brand[b].append(p)
        except Exception:
            pass

    all_products = []
    for i, brand in enumerate(brands):
        # Check cancel signal
        if crawl_cancel_event.is_set():
            crawl_progress["message"] = "Curación cancelada"
            print("🛑 Crawl cancelled by user")
            break

        crawl_progress["brand_idx"] = i + 1
        crawl_progress["current_brand"] = brand["name"]
        crawl_progress["message"] = f"Crawleando {brand['name']}... ({i+1}/{len(brands)})"

        print(f"Crawling {brand['name']}...")

        def progress_cb(msg):
            crawl_progress["message"] = msg

        # Run crawl_brand with a 3-minute timeout per brand
        brand_result = [None]
        brand_error = [None]
        def _crawl_one():
            try:
                brand_result[0] = crawl_brand(brand, progress_callback=progress_cb)
            except Exception as e:
                brand_error[0] = e

        brand_thread = threading.Thread(target=_crawl_one, daemon=True)
        brand_thread.start()
        brand_thread.join(timeout=180)  # 3 minutes max per brand

        if brand_thread.is_alive():
            print(f"  ⏱ TIMEOUT: {brand['name']} took >3 min, skipping")
            crawl_progress["message"] = f"⏱ {brand['name']}: timeout (>3 min)"
            crawl_progress["failed_brands"].append(brand["name"])
            # Thread is daemon, will die when main exits
            prev = prev_products_by_brand.get(brand["name"], [])
            if prev:
                all_products.extend(prev)
            time.sleep(1)
            continue

        if brand_error[0]:
            print(f"  ❌ {brand['name']} error: {brand_error[0]}")
            crawl_progress["failed_brands"].append(brand["name"])
            prev = prev_products_by_brand.get(brand["name"], [])
            if prev:
                all_products.extend(prev)
            time.sleep(1)
            continue

        products = brand_result[0]

        if not products:
            # PRESERVE previous data for this brand if crawl fails
            prev = prev_products_by_brand.get(brand["name"], [])
            if prev:
                all_products.extend(prev)
                crawl_progress["message"] = f"⚠ {brand['name']}: sin acceso (usando {len(prev)} del cache anterior)"
                print(f"  → 0 new products, preserved {len(prev)} from previous cache")
            else:
                crawl_progress["message"] = f"⚠ {brand['name']}: sin acceso"
                print(f"  → 0 products (failed)")
            crawl_progress["failed_brands"].append(brand["name"])
            time.sleep(2.0)
            continue

        products = filter_unwanted_products(products)
        before = len(products)
        products = deduplicate_products(products)
        after = len(products)
        all_products.extend(products)

        crawl_progress["products_found"] = len(all_products)
        crawl_progress["message"] = f"✓ {brand['name']}: {after} productos"

        if before != after:
            print(f"  → {before} products, {before - after} dupes removed → {after}")
        else:
            print(f"  → {after} products")

        # Guardar cache parcial después de cada marca
        try:
            partial = {"products": all_products, "crawled_at": datetime.now().isoformat(), "partial": True}
            tmp_path = cache_file + ".tmp"
            with open(tmp_path, "w") as f:
                json.dump(partial, f, ensure_ascii=False)
            os.replace(tmp_path, cache_file)
        except Exception as e:
            print(f"  ⚠ Partial cache save failed: {e}")

        # Esperar entre marcas
        if i < len(brands) - 1:
            time.sleep(3.0)

    # Dedup final entre marcas
    total_before = len(all_products)
    all_products = deduplicate_products(all_products)
    if total_before != len(all_products):
        print(f"Cross-brand dedup: {total_before} → {len(all_products)}")

    # Guardar cache final
    cache_data = {"products": all_products, "crawled_at": datetime.now().isoformat()}
    tmp_path = cache_file + ".tmp"
    with open(tmp_path, "w") as f:
        json.dump(cache_data, f, ensure_ascii=False)
        f.flush()
        os.fsync(f.fileno())
    os.replace(tmp_path, cache_file)
    print(f"Cache saved: {len(all_products)} products → {cache_file}")
    # También guardar en Firestore para persistencia
    if country:
        save_cache_firestore(all_products, country)
    else:
        print("⚠️ No country set — cache not saved to Firestore")

    crawl_progress["status"] = "done"
    crawl_progress["done"] = True
    crawl_progress["products_found"] = len(all_products)
    crawl_progress["message"] = f"✅ Listo: {len(all_products)} productos de {len(brands)} marcas"

    # Log crawl to Firestore for audit trail
    try:
        db = _get_db_safe()
        if db:
            db.collection("curator").document("_crawl_log").collection("runs").add({
                "user": user_id or "unknown",
                "country": country or "",
                "brands_total": len(brands),
                "brands_failed": crawl_progress.get("failed_brands", []),
                "products_found": len(all_products),
                "timestamp": firestore_timestamp(),
                "duration_brands": {b["name"]: "ok" for b in brands if b["name"] not in crawl_progress.get("failed_brands", [])},
            })
    except Exception as e:
        print(f"⚠️ Error logging crawl: {e}")

    return all_products


def load_crawl_cache():
    if os.path.exists(CRAWL_CACHE):
        with open(CRAWL_CACHE) as f:
            data = json.load(f)
            return data.get("products", [])
    return None


# ─── Excel generation ────────────────────────────────────────────────────

def _norm_url(url):
    """Normalize URL for comparison: strip trailing slash, lowercase."""
    return url.rstrip("/").lower() if url else ""

def _sanitize_cell(value):
    """Prevent Excel formula injection (CWE-1236). Prefix dangerous chars with apostrophe."""
    if isinstance(value, str) and value and value[0] in ('=', '+', '-', '@', '\t', '\r'):
        return "'" + value
    return value

def build_curated_brands_for_ordering(session, country, brand_selection_order=None):
    """Cruza session['accepted'] + session['previous_rows'] con el crawl cache
    para reconstruir las marcas aprobadas con toda la información necesaria
    para la UI del Paso 4 "Ordenar y Descargar" y para el upload directo al admin.

    Fuentes de datos (merge policy):

    1. **previous_rows**: filas de la planilla .xlsx que el admin subió al inicio
       de la sesión. Ya vienen con Orden/Posición/Etiquetas asignadas. Solo
       incluimos las filas con `Aprobado == "Si"` (las rechazadas se omiten
       del visualizador pero se preservan en `previous_rows` por si el admin
       descarga el Excel directo sin Paso 4).

    2. **accepted**: productos que el admin aprobó/marcó como trend en esta
       sesión de curación. Cruzan con `_get_cached_products(country)` para
       obtener imagen/categoría/tags del crawl.

    **Política de duplicados**: si un URL aparece en ambas fuentes (el admin
    re-curó un producto que ya estaba en la planilla), **gana la versión nueva**
    (accepted). La decisión se confirmó con el usuario en la auditoría
    del 12 abril 2026: el re-curado implica intención de actualizar.

    Returns:
        list of dicts, each: {
            "name": brand_name,
            "products": [
                {
                    "url": str,
                    "title": str,
                    "image_url": str,
                    "category": str,
                    "tags": list[str],         # hasta 4 etiquetas
                    "is_trend": bool,
                    "source": "accepted" | "previous",   # para debugging
                },
                ...
            ]
        }

    Args:
        session: el dict de session del usuario (load_session())
        country: código ISO del país activo
        brand_selection_order: lista ordenada de nombres de marca (opcional).
            Si se provee, las marcas en el resultado respetan ese orden.
            Si no, se usa el orden de primera aparición combinando ambas
            fuentes (previous_rows primero porque son los "anclas" de la
            planilla cargada, accepted después).
    """
    all_products = _get_cached_products(country) or []

    # Mapa url_normalizada → producto crawleado (para O(1) lookup en accepted)
    by_url = {}
    for p in all_products:
        by_url[_norm_url(p.get("product_url", ""))] = p

    # ── PASO 1: accepted set (se procesa primero para tener prioridad) ──
    accepted_info = {}  # url → {"is_trend": bool}
    for p in session.get("accepted", []):
        url = _norm_url(p.get("product_url", ""))
        accepted_info[url] = {"is_trend": bool(p.get("trend"))}

    # Build un dict unificado: url_norm → {brand, product_dict}
    # Si la misma url aparece en ambas fuentes, la versión 'accepted' gana.
    merged_by_url = {}
    # Preservamos el orden de aparición de las marcas para el fallback
    # brand_first_seen cuando no hay brand_selection_order.
    brand_first_seen_order = []
    seen_brands = set()

    # ── PASO 2: PREVIOUS_ROWS primero (se van a sobrescribir por accepted
    #           si hay conflicto). Sólo incluimos filas con Aprobado == "Si".
    for prev_row in session.get("previous_rows", []) or []:
        link = prev_row.get("Link", "")
        if not isinstance(link, str) or not link.startswith("http"):
            continue
        aprobado = str(prev_row.get("Aprobado", "No")).strip().lower()
        # Decisión del usuario (auditoría 12 abril 2026): omitir rechazados
        # del visualizador del Paso 4. Siguen existiendo en previous_rows
        # y se preservan al descargar el Excel directo.
        if aprobado not in ("si", "sí", "yes", "true"):
            continue

        url_norm = _norm_url(link)
        brand = str(prev_row.get("Marca", "Desconocida")).strip() or "Desconocida"

        # Extraer etiquetas de las columnas del Excel (pueden estar vacías)
        prev_tags = [
            str(prev_row.get("Etiqueta 1", "") or "").strip(),
            str(prev_row.get("Etiqueta 2", "") or "").strip(),
            str(prev_row.get("Etiqueta 3", "") or "").strip(),
            str(prev_row.get("Etiqueta 4", "") or "").strip(),
        ]
        prev_tags = [t for t in prev_tags if t]  # drop vacíos

        is_trend = str(prev_row.get("Tendencia", "No")).strip().lower() in ("si", "sí", "yes", "true")

        # Intentar obtener imagen del crawl cache si existe.
        # Si no, fallback a la columna Imagen del Excel (puede estar vacía).
        cached = by_url.get(url_norm)
        image_url = ""
        title = ""
        category = str(prev_row.get("Categoría", "") or "").strip()
        if cached:
            image_url = cached.get("image_url", "") or ""
            title = cached.get("name", "") or cached.get("title", "") or ""
        # Fallback a columna Imagen/Titulo del Excel (si el admin la llenó)
        if not image_url:
            excel_img = prev_row.get("Imagen", "")
            if isinstance(excel_img, str) and excel_img.startswith("http"):
                image_url = excel_img
        if not title:
            excel_title = prev_row.get("Titulo", "") or prev_row.get("Título", "")
            if isinstance(excel_title, str):
                title = str(excel_title).strip()
        # Último recurso: derivar un título legible del URL (el slug de la
        # última parte del path). Mejor que un card sin texto en el Paso 4.
        if not title:
            try:
                from urllib.parse import urlparse
                path = urlparse(link).path
                slug = path.rstrip("/").split("/")[-1] or "(sin título)"
                # Capitalizar y reemplazar guiones/underscores por espacios
                title = slug.replace("-", " ").replace("_", " ").strip().title()
                if not title:
                    title = "(sin título)"
            except Exception:
                title = "(sin título)"

        merged_by_url[url_norm] = {
            "brand": brand,
            "url": link,
            "title": title,
            "image_url": image_url,
            "category": category,
            "tags": prev_tags[:4],
            "is_trend": is_trend,
            "source": "previous",
        }

        if brand not in seen_brands:
            brand_first_seen_order.append(brand)
            seen_brands.add(brand)

    # ── PASO 3: ACCEPTED — sobrescribe si hay conflicto por URL ──
    for p in all_products:
        url_norm = _norm_url(p.get("product_url", ""))
        if url_norm not in accepted_info:
            continue  # Solo productos aprobados esta sesión

        brand = p.get("brand", "Desconocida")
        tags = p.get("tags", [])
        if not isinstance(tags, list):
            tags = []

        merged_by_url[url_norm] = {
            "brand": brand,
            "url": p.get("product_url", ""),
            "title": p.get("name", "") or p.get("title", ""),
            "image_url": p.get("image_url", ""),
            "category": p.get("category", ""),
            "tags": [_sanitize_cell(t) for t in tags[:4]],
            "is_trend": accepted_info[url_norm]["is_trend"],
            "source": "accepted",
        }

        if brand not in seen_brands:
            brand_first_seen_order.append(brand)
            seen_brands.add(brand)

    # ── PASO 4: Agrupar por marca ──
    brand_groups = {}
    for url_norm, item in merged_by_url.items():
        brand = item["brand"]
        if brand not in brand_groups:
            brand_groups[brand] = []
        # Remove the `brand` key; downstream expects the product dict shape.
        prod_copy = {k: v for k, v in item.items() if k != "brand"}
        brand_groups[brand].append(prod_copy)

    # ── PASO 5: Determinar orden final de marcas ──
    if brand_selection_order:
        ordered_names = [b for b in brand_selection_order if b in brand_groups]
        extras = [b for b in brand_first_seen_order if b not in brand_selection_order and b in brand_groups]
        ordered_names += extras
    else:
        ordered_names = [b for b in brand_first_seen_order if b in brand_groups]

    return [{"name": name, "products": brand_groups[name]} for name in ordered_names]


def generate_plantilla(all_products, accepted_urls, trend_urls, output_path,
                       previous_rows=None, brand_order=None, product_order_override=None):
    """Generate moder_plantilla_productos.xlsx — ALL products with Aprobado Si/No column.

    Contrato de columnas (semántica unificada robot ↔ admin ↔ iOS):
      - Orden      = importancia de la MARCA (1 = más importante). Todos los
                     productos de una misma marca comparten el mismo valor de Orden.
                     Mapea a `stores/{id}.order` en Firestore y `StoreModel.order`
                     en iOS.
      - Posición   = importancia del PRODUCTO dentro de su marca (1..N). Mapea a
                     `stores/{id}.products[i].order` en Firestore y
                     `ProductLink.order` en iOS.
      - Etiqueta 4 = cuarta etiqueta opcional para filtros de usuaria. Mapea a
                     `products[i].tag4` en Firestore y `ProductLink.tag4` en iOS.

    Args:
        all_products: list of all crawled product dicts
        accepted_urls: set of URLs that were approved
        trend_urls: set of URLs marked as trending
        output_path: file path for the xlsx
        previous_rows: list of dicts from a previously uploaded spreadsheet
        brand_order: optional list of brand names in the desired order of
                     importance (most important first). If None, uses the
                     natural order of first-appearance in `all_products`
                     (preserves robot's active_brands selection order).
        product_order_override: optional dict {brand_name: [url1, url2, ...]}
                     that defines the order of products WITHIN each brand.
                     If None, uses the natural order as products were crawled.
                     This is populated by Paso 4 "Ordenar y Descargar".
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    headers = ["Link", "Marca", "Aprobado", "Tendencia", "Orden", "Posición",
               "Top 20", "Categoría", "Etiqueta 1", "Etiqueta 2", "Etiqueta 3", "Etiqueta 4"]

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1C1C1E", end_color="1C1C1E", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    existing_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    approved_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    rejected_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

    row = 2

    # Detectar modo Paso 4: si hay brand_order explícito + product_order_override,
    # el usuario pasó por el visualizador del Paso 4 y reordenó las marcas y
    # productos (incluyendo las filas de la planilla previa). En ese modo,
    # necesitamos "promover" las previous_rows aprobadas al flujo principal de
    # reordenamiento, en lugar de escribirlas "as-is" al inicio.
    step4_mode = bool(brand_order or product_order_override)

    # ── PART 1: Write previous rows ──
    prev_urls = set()
    # En modo Paso 4: guardamos las filas previas aprobadas para inyectarlas
    # en brand_groups más abajo. Las rechazadas sí van al Excel as-is al final
    # para preservar el historial (por si el admin quiere reconsiderarlas).
    promoted_prev_products = []  # list of (brand, synthetic_product_dict)
    rejected_prev_rows = []      # se escriben al final, sin orden del Paso 4

    if previous_rows:
        for prev_row in previous_rows:
            link = str(prev_row.get("Link", ""))
            if not link.startswith("http"):
                continue
            prev_urls.add(_norm_url(link))

            if step4_mode:
                # Clasificar: aprobado va al reorder, rechazado queda aparte.
                aprobado = str(prev_row.get("Aprobado", "No")).strip().lower()
                if aprobado in ("si", "sí", "yes", "true"):
                    # Construir producto sintético compatible con all_products
                    # shape. Lo agregamos al brand_groups después.
                    brand = str(prev_row.get("Marca", "Desconocida")).strip() or "Desconocida"
                    tags_list = [
                        str(prev_row.get("Etiqueta 1", "") or "").strip(),
                        str(prev_row.get("Etiqueta 2", "") or "").strip(),
                        str(prev_row.get("Etiqueta 3", "") or "").strip(),
                        str(prev_row.get("Etiqueta 4", "") or "").strip(),
                    ]
                    synthetic = {
                        "product_url": link,
                        "brand": brand,
                        "name": prev_row.get("Titulo", "") or prev_row.get("Título", "") or "",
                        "category": str(prev_row.get("Categoría", "") or "").strip(),
                        "tags": [t for t in tags_list if t],
                        "image_url": prev_row.get("Imagen", "") if isinstance(prev_row.get("Imagen", ""), str) else "",
                        "_from_previous": True,  # marker para distinguir en el loop
                        "_prev_is_trend": str(prev_row.get("Tendencia", "No")).strip().lower() in ("si", "sí", "yes", "true"),
                    }
                    promoted_prev_products.append((brand, synthetic))
                else:
                    rejected_prev_rows.append(prev_row)
                continue  # no escribir as-is en modo Paso 4

            # Modo legacy: escribir as-is tal como estaba antes.
            col_map = {
                1: link,
                2: prev_row.get("Marca", ""),
                3: prev_row.get("Aprobado", prev_row.get("Top 20", "No")),
                4: prev_row.get("Tendencia", "No"),
                5: prev_row.get("Orden", ""),
                6: prev_row.get("Posición", ""),
                7: prev_row.get("Top 20", "No"),
                8: prev_row.get("Categoría", ""),
                9: prev_row.get("Etiqueta 1", ""),
                10: prev_row.get("Etiqueta 2", ""),
                11: prev_row.get("Etiqueta 3", ""),
                12: prev_row.get("Etiqueta 4", ""),
            }
            for c, val in col_map.items():
                cell = ws.cell(row=row, column=c, value=val if val else "")
                cell.border = thin_border
                cell.fill = existing_fill

            if link:
                ws.cell(row=row, column=1).hyperlink = link
                ws.cell(row=row, column=1).font = Font(color="666666", underline="single")

            row += 1

    # ── PART 2: Write ALL new products with Aprobado Si/No ──
    # En modo Paso 4 incluimos también los "promoted_prev_products" que venían
    # de previous_rows aprobadas, para que entren al reordenamiento con los overrides.
    brand_groups = {}
    brand_first_seen_order = []

    # Primero los sintéticos de previous_rows aprobadas (modo Paso 4).
    # Así respetan el orden de aparición para el fallback si no hay brand_order.
    for brand, synthetic in promoted_prev_products:
        if brand not in brand_groups:
            brand_groups[brand] = []
            brand_first_seen_order.append(brand)
        brand_groups[brand].append(synthetic)
        # Marcar la URL como ya escrita en "previous" para evitar doble-write
        # desde all_products en el próximo loop.
        # (prev_urls ya tiene estas URLs del loop anterior.)

    # Ahora los productos del crawl que NO están en previous.
    for p in all_products:
        url = _norm_url(p.get("product_url", ""))
        if url in prev_urls:
            continue  # Already handled (as-is or promoted al step4 mode)
        brand = p.get("brand", "Desconocida")
        if brand not in brand_groups:
            brand_groups[brand] = []
            brand_first_seen_order.append(brand)
        brand_groups[brand].append(p)

    # Paso 2: determinar el orden FINAL de las marcas.
    # Prioridad:
    #   1. brand_order explícito (viene del Paso 4 "Ordenar y descargar")
    #   2. orden de primera aparición en all_products (respeta active_brands)
    if brand_order:
        # Filtrar solo marcas que realmente tienen productos, preservando el
        # orden indicado. Marcas extra (no en brand_order) van al final.
        ordered_brand_names = [b for b in brand_order if b in brand_groups]
        extras = [b for b in brand_first_seen_order if b not in brand_order]
        ordered_brand_names += extras
    else:
        ordered_brand_names = brand_first_seen_order

    new_count = 0
    # brand_idx = 1-based rank of the brand in the ordered list (Column "Orden")
    for brand_idx, brand_name in enumerate(ordered_brand_names, start=1):
        products = brand_groups[brand_name]

        # Si el Paso 4 pasó un orden de productos dentro de la marca, úsalo.
        # Si no, preservar orden natural de crawl.
        if product_order_override and brand_name in product_order_override:
            url_order = product_order_override[brand_name]
            # Sort products by index in url_order; unseen URLs go to the end
            def _rank(p):
                u = _norm_url(p.get("product_url", ""))
                try:
                    return url_order.index(u)
                except ValueError:
                    return len(url_order) + 9999
            products = sorted(products, key=_rank)

        for position_in_brand, p in enumerate(products, start=1):
            url = _norm_url(p.get("product_url", ""))

            # Detectar si el producto viene de previous_rows (promoted al
            # modo Paso 4). En ese caso, is_approved/is_trend vienen del
            # dict sintético — no de los sets accepted/trend_urls que solo
            # reflejan la sesión actual de curación.
            is_from_previous = p.get("_from_previous", False)
            if is_from_previous:
                is_approved = True  # promoted solo se activa si "Aprobado": "Si"
                is_trend = bool(p.get("_prev_is_trend", False))
            else:
                is_approved = url in accepted_urls
                is_trend = url in trend_urls

            tags = p.get("tags", [])
            if not isinstance(tags, list):
                tags = []

            ws.cell(row=row, column=1, value=_sanitize_cell(p.get("product_url", ""))).border = thin_border
            ws.cell(row=row, column=2, value=_sanitize_cell(brand_name)).border = thin_border
            ws.cell(row=row, column=3, value="Si" if is_approved else "No").border = thin_border
            ws.cell(row=row, column=4, value="Si" if is_trend else "No").border = thin_border
            # Orden = rank de la MARCA (igual para todos los productos de la marca)
            ws.cell(row=row, column=5, value=brand_idx).border = thin_border
            # Posición = rank del PRODUCTO dentro de su marca (1..N)
            ws.cell(row=row, column=6, value=position_in_brand).border = thin_border
            ws.cell(row=row, column=7, value="No").border = thin_border
            ws.cell(row=row, column=8, value=_sanitize_cell(p.get("category", ""))).border = thin_border
            ws.cell(row=row, column=9, value=_sanitize_cell(tags[0]) if len(tags) > 0 else "").border = thin_border
            ws.cell(row=row, column=10, value=_sanitize_cell(tags[1]) if len(tags) > 1 else "").border = thin_border
            ws.cell(row=row, column=11, value=_sanitize_cell(tags[2]) if len(tags) > 2 else "").border = thin_border
            # Etiqueta 4 — opcional, se llena en el Paso 4 o manualmente.
            ws.cell(row=row, column=12, value=_sanitize_cell(tags[3]) if len(tags) > 3 else "").border = thin_border

            # Color: green for approved, orange for rejected
            fill = approved_fill if is_approved else rejected_fill
            for c in range(1, 13):
                ws.cell(row=row, column=c).fill = fill

            ws.cell(row=row, column=1).hyperlink = p.get("product_url", "")
            ws.cell(row=row, column=1).font = Font(
                color="0066CC" if is_approved else "999999", underline="single")

            row += 1
            new_count += 1

    # ── PART 3: Write rejected previous rows at the end (modo Paso 4 only) ──
    # Los rechazados de la planilla previa no entran al reorder del Paso 4,
    # pero los escribimos al final del Excel como archivo histórico para
    # que el admin pueda verlos si quiere reconsiderarlos más adelante.
    if step4_mode and rejected_prev_rows:
        for prev_row in rejected_prev_rows:
            link = str(prev_row.get("Link", ""))
            if not link.startswith("http"):
                continue
            col_map = {
                1: link,
                2: prev_row.get("Marca", ""),
                3: prev_row.get("Aprobado", "No"),
                4: prev_row.get("Tendencia", "No"),
                5: prev_row.get("Orden", ""),
                6: prev_row.get("Posición", ""),
                7: prev_row.get("Top 20", "No"),
                8: prev_row.get("Categoría", ""),
                9: prev_row.get("Etiqueta 1", ""),
                10: prev_row.get("Etiqueta 2", ""),
                11: prev_row.get("Etiqueta 3", ""),
                12: prev_row.get("Etiqueta 4", ""),
            }
            for c, val in col_map.items():
                cell = ws.cell(row=row, column=c, value=val if val else "")
                cell.border = thin_border
                cell.fill = existing_fill
            ws.cell(row=row, column=1).hyperlink = link
            ws.cell(row=row, column=1).font = Font(color="666666", underline="single")
            row += 1

    # Column widths — agregamos columna 12 (Etiqueta 4)
    widths = [55, 22, 10, 10, 8, 10, 8, 15, 15, 15, 15, 15]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    ws.auto_filter.ref = f"A1:L{row - 1}"
    ws.freeze_panes = "A2"

    # Save to disk
    wb.save(output_path)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    approved_count = sum(1 for p in all_products if _norm_url(p.get("product_url", "")) in accepted_urls)
    print(f"Plantilla: {len(previous_rows or [])} anteriores + {new_count} nuevos ({approved_count} aprobados) = {row - 2} total")
    return output_path, buf


# ─── Parse previous spreadsheet ─────────────────────────────────────────

def parse_previous_spreadsheet(file_path):
    """Extract URLs and full row data from a previously uploaded spreadsheet"""
    urls = set()
    rows_data = []  # Full row data for accumulation
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active

        # Read headers
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
            headers.append(str(cell) if cell else "")

        # Find the Link column by header name (not position)
        link_col = 0
        for i, h in enumerate(headers):
            if h.lower().strip() in ("link", "url", "enlace"):
                link_col = i
                break

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            # Get URL from the Link column
            link_val = row[link_col] if link_col < len(row) else None
            if link_val and isinstance(link_val, str) and link_val.startswith("http"):
                url = _norm_url(link_val.strip())
                urls.add(url)
                # Store full row as dict
                row_dict = {}
                for i, val in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = val if val else ""
                rows_data.append(row_dict)
    except Exception as e:
        print(f"Error parsing spreadsheet: {e}")
    return urls, rows_data


# ─── Auth Routes ─────────────────────────────────────────────────────────

# ─── Health Check & Monitoring ─────────────────────────────────────────

@app.route("/health")
def health_check():
    """Health check for Render. Basic status for unauthenticated, details for logged-in."""
    import time as _t
    is_admin = flask_session.get("logged_in", False)
    status = {"status": "ok"}

    # Firestore connectivity (always check — affects Render health)
    try:
        db = _get_db_safe()
        if db:
            db.collection("curator").document("_healthcheck").set({
                "last_ping": datetime.now().isoformat()
            })
        else:
            status["status"] = "degraded"
    except Exception:
        status["status"] = "degraded"

    # Detailed diagnostics only for authenticated users
    if is_admin:
        status["timestamp"] = datetime.now().isoformat()
        status["firestore"] = "connected" if status["status"] == "ok" else "degraded"
        try:
            import resource
            mem_mb = resource.getrusage(resource.RUSAGE_SELF).ru_maxrss / 1024 / 1024
            status["memory_mb"] = round(mem_mb, 1)
        except Exception:
            pass
        locked = not crawl_lock.acquire(blocking=False)
        if not locked:
            crawl_lock.release()
        status["crawl_locked"] = locked
        status["cached_sessions"] = len(_session_cache)
        status["cached_xlsx"] = len(generated_xlsx_per_user)

    # Periodic cleanup: evict stale local files older than 7 days (run max once per hour)
    global _last_cleanup_time
    now = _t.time()
    if now - _last_cleanup_time > 3600:
        _last_cleanup_time = now
        try:
            import glob as _glob
            cutoff = now - (7 * 86400)
            stale = 0
            for pattern in ["session_*.json", "crawl_cache_*.json"]:
                for f in _glob.glob(os.path.join(DATA_DIR, pattern)):
                    if os.path.getmtime(f) < cutoff:
                        os.remove(f)
                        stale += 1
            if stale:
                status["stale_files_cleaned"] = stale
        except Exception:
            pass

    code = 200 if status["status"] == "ok" else 503
    return jsonify(status), code


def log_error_to_firestore(error_type, message, details=None):
    """Log errors to Firestore for visibility without Sentry."""
    try:
        db = _get_db_safe()
        if not db:
            return
        db.collection("curator").document("_errors").collection("log").add({
            "type": error_type,
            "message": str(message)[:500],
            "details": str(details)[:1000] if details else "",
            "timestamp": firestore_timestamp(),
            "user": _safe_get_user_id(),
        })
    except Exception:
        pass  # Don't let error logging cause more errors


@app.route("/login")
def login_page():
    if flask_session.get("logged_in"):
        return redirect(url_for("index"))
    return render_template("login.html")


@app.route("/login", methods=["POST"])
def login_action():
    data = request.get_json(silent=True) or {}
    username = data.get("username", "").strip()
    password = data.get("password", "")

    if username in USERS and check_password_hash(USERS[username], password):
        flask_session["logged_in"] = True
        flask_session["username"] = username
        # Clear in-memory session cache so fresh data is loaded
        invalidate_session_cache(username)
        return jsonify({"status": "ok"})
    return jsonify({"status": "error", "error": "Usuario o contraseña incorrectos"}), 401


@app.route("/logout")
def logout():
    flask_session.clear()
    return redirect(url_for("login_page"))


# ─── Routes ──────────────────────────────────────────────────────────────

@app.route("/select-country", methods=["POST"])
@login_required
def select_country():
    """Set the active country for this session"""
    data = request.get_json(silent=True) or {}
    country_code = data.get("country", "").upper()
    if country_code not in COUNTRIES:
        return jsonify({"error": "Invalid country"}), 400
    save_active_country(country_code)
    return jsonify({"status": "ok", "country": country_code})


@app.route("/")
@login_required
def index():
    country = load_active_country()
    if not country:
        return render_template("index.html",
                               selecting_country=True,
                               countries=COUNTRIES,
                               has_cache=False, product_count=0,
                               accepted_count=0, rejected_count=0,
                               previous_count=0, active_brands=[],
                               suggested_brands=[], default_brands=[],
                               hidden_count=0,
                               active_country="", country_info={})
    session = load_session()
    # Load country-specific cache (Firestore → local file fallback)
    cache_path = get_cache_file_for_country(country)
    products = None
    if os.path.exists(cache_path):
        try:
            with open(cache_path) as f:
                data = json.load(f)
                products = data.get("products", [])
        except (json.JSONDecodeError, IOError):
            pass
    # Fallback a Firestore si archivo local no existe
    if not products:
        fs_products = load_cache_firestore(country)
        if fs_products:
            products = fs_products
    active_brands = load_active_brands(country)
    has_cache = products is not None and len(products) > 0
    all_suggested = SUGGESTED_BRANDS_BY_COUNTRY.get(country, [])
    active_domains = set(b["domain"] for b in active_brands)
    hidden_domains = set(load_hidden_brands(country))
    suggested = [b for b in all_suggested if b["domain"] not in active_domains and b["domain"] not in hidden_domains]
    return render_template("index.html",
                           selecting_country=False,
                           countries=COUNTRIES,
                           has_cache=has_cache,
                           product_count=len(products) if products else 0,
                           accepted_count=len(session.get("accepted", [])),
                           rejected_count=len(session.get("rejected", [])),
                           previous_count=len(session.get("previous_urls", [])),
                           active_brands=active_brands,
                           suggested_brands=suggested,
                           hidden_count=len(hidden_domains),
                           default_brands=DEFAULT_BRANDS,
                           active_country=country,
                           country_info=COUNTRIES.get(country, {}))


@app.route("/update-brands", methods=["POST"])
@login_required
def update_brands():
    """Update active brand list"""
    country = load_active_country()
    data = request.get_json(silent=True) or {}
    brands = data.get("brands", [])
    save_active_brands(brands, country)
    return jsonify({"status": "ok", "count": len(brands)})


@app.route("/add-all-suggested", methods=["POST"])
@login_required
def add_all_suggested():
    """Add all suggested brands at once for current country"""
    country = load_active_country()
    active = load_active_brands(country)
    active_domains = set(b["domain"] for b in active)
    suggested = SUGGESTED_BRANDS_BY_COUNTRY.get(country, [])
    added = 0
    for brand in suggested:
        if brand["domain"] not in active_domains:
            active.append(brand)
            active_domains.add(brand["domain"])
            added += 1
    save_active_brands(active, country)
    return jsonify({"status": "ok", "added": added, "total": len(active)})


@app.route("/add-brand", methods=["POST"])
@login_required
def add_brand():
    """Add a custom brand by URL"""
    country = load_active_country()
    data = request.get_json(silent=True) or {}
    name = data.get("name", "").strip().upper()
    url = data.get("url", "").strip()

    if not name or not url:
        return jsonify({"error": "Name and URL required"}), 400

    # Ensure URL has protocol
    if not url.startswith("http"):
        url = "https://" + url

    # Extract domain
    from urllib.parse import urlparse
    domain = urlparse(url).netloc.replace("www.", "")

    active = load_active_brands(country)
    # Check not already present
    if any(b["domain"] == domain for b in active):
        return jsonify({"error": "Brand already exists"}), 400

    new_brand = {"name": name, "domain": domain, "url": url}
    active.append(new_brand)
    save_active_brands(active, country)

    return jsonify({"status": "ok", "brand": new_brand})


@app.route("/remove-brand", methods=["POST"])
@login_required
def remove_brand():
    """Remove a brand from active list"""
    country = load_active_country()
    data = request.get_json(silent=True) or {}
    domain = data.get("domain", "")
    active = load_active_brands(country)
    active = [b for b in active if b["domain"] != domain]
    save_active_brands(active, country)
    return jsonify({"status": "ok", "count": len(active)})


@app.route("/remove-all-brands", methods=["POST"])
@login_required
def remove_all_brands():
    """Remove all brands from active list and clear curation session"""
    country = load_active_country()
    save_active_brands([], country)
    # Clear curation session and cache
    clear_curation_session(country)
    return jsonify({"status": "ok", "count": 0})


@app.route("/hide-brand", methods=["POST"])
@login_required
def hide_brand():
    """Permanently hide a brand from the suggested list"""
    country = load_active_country()
    data = request.get_json(silent=True) or {}
    domain = data.get("domain", "")
    if not domain:
        return jsonify({"error": "domain required"}), 400
    hidden = load_hidden_brands(country)
    if domain not in hidden:
        hidden.append(domain)
        save_hidden_brands(hidden, country)
    # Also remove from active if present
    active = load_active_brands(country)
    active = [b for b in active if b["domain"] != domain]
    save_active_brands(active, country)
    return jsonify({"status": "ok", "hidden_count": len(hidden)})


@app.route("/unhide-brands", methods=["POST"])
@login_required
def unhide_brands():
    """Restore all hidden brands back to suggested list"""
    country = load_active_country()
    save_hidden_brands([], country)
    return jsonify({"status": "ok"})


@app.route("/change-country", methods=["POST"])
@login_required
def change_country():
    """Switch to a different country"""
    save_active_country("")
    return jsonify({"status": "ok"})


@app.route("/crawl", methods=["POST"])
@login_required
def crawl():
    """Start crawling in background thread"""
    country = load_active_country()
    active_brands = load_active_brands(country)
    if not active_brands:
        return jsonify({"error": "No brands selected"}), 400

    # Prevent concurrent crawls — auto-expire lock after 5 minutes
    import time as _time
    global crawl_lock_time
    if not crawl_lock.acquire(blocking=False):
        # Check if lock is stale (>10 min = crashed thread)
        if _time.time() - crawl_lock_time > 300:  # 5 min auto-expire
            print("⚠️ Crawl lock stale (>10 min), force-releasing...")
            try:
                crawl_lock.release()
            except RuntimeError:
                pass
            crawl_lock.acquire(blocking=False)
        else:
            return jsonify({"error": "Ya hay un crawl en progreso. Espera a que termine."}), 429
    crawl_lock_time = _time.time()

    # Reset curation index but PRESERVE accepted/rejected if user has data
    session = load_session()
    session["current_index"] = 0
    # Only clear if explicitly requested (not by default — protects user data)
    req_data = request.get_json(silent=True) or {}
    if req_data.get("clear_session"):
        session["accepted"] = []
        session["rejected"] = []
    save_session(session)

    # Use country-specific cache file
    cache_file = get_cache_file_for_country(country)
    uid = get_user_id()
    progress = get_crawl_progress(uid)

    def run_crawl():
        try:
            print(f"🚀 Starting crawl for {len(active_brands)} brands in {country} (user: {uid})...")
            products = crawl_all(active_brands, cache_file=cache_file, progress=progress, country=country, user_id=uid)
            print(f"✅ Crawl complete: {len(products)} products from {len(active_brands)} brands")
        except Exception as e:
            print(f"❌ CRAWL ERROR: {e}")
            import traceback
            traceback.print_exc()
            progress["done"] = True
            progress["status"] = "error"
            progress["message"] = f"Error: {str(e)[:100]}"
        finally:
            crawl_lock.release()

    t = threading.Thread(target=run_crawl, daemon=True)
    t.start()
    return jsonify({"status": "started", "brands": len(active_brands)})


@app.route("/force-unlock", methods=["POST"])
@login_required
def force_unlock():
    """Force-release the crawl lock if it's stuck. Only admin (sebastian) can force-unlock."""
    if get_user_id() != "sebastian":
        return jsonify({"error": "Solo el administrador puede desbloquear"}), 403
    try:
        crawl_lock.release()
    except RuntimeError:
        pass  # Already unlocked
    uid = get_user_id()
    progress = get_crawl_progress(uid)
    progress["done"] = True
    progress["status"] = "cancelled"
    progress["message"] = "Crawl desbloqueado manualmente"
    return jsonify({"status": "ok"})


@app.route("/crawl-progress")
@login_required
def crawl_progress_endpoint():
    """Polling endpoint — returns current user's crawl progress"""
    uid = get_user_id()
    return jsonify(get_crawl_progress(uid))


@app.route("/upload-previous", methods=["POST"])
@login_required
def upload_previous():
    """Upload previous spreadsheet — becomes the accumulation base"""
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file"}), 400

    # Validate file size (max 10MB)
    file.seek(0, 2)
    size = file.tell()
    file.seek(0)
    if size > 10 * 1024 * 1024:
        return jsonify({"error": "Archivo demasiado grande (máx 10MB)"}), 400
    if not file.filename.endswith('.xlsx'):
        return jsonify({"error": "Solo se aceptan archivos .xlsx"}), 400

    uid = get_user_id()
    path = os.path.join(DATA_DIR, f"previous_upload_{uid}.xlsx")
    file.save(path)
    urls, rows_data = parse_previous_spreadsheet(path)

    session = load_session()
    session["previous_urls"] = list(urls)
    session["previous_rows"] = rows_data  # Full row data for accumulation
    save_session(session)

    return jsonify({"status": "ok", "known_urls": len(urls), "known_products": len(rows_data)})


@app.route("/curate")
@login_required
def curate():
    """Main curation interface"""
    country = load_active_country()
    session = load_session()
    # Load country-specific cache
    cache_path = get_cache_file_for_country(country)
    products = None
    if os.path.exists(cache_path):
        try:
            with open(cache_path) as f:
                data = json.load(f)
                products = data.get("products", [])
        except (json.JSONDecodeError, IOError):
            pass
    if not products:
        # Fallback to Firestore cache (no legacy global — prevents country leak)
        fs_products = load_cache_firestore(country)
        if fs_products:
            products = fs_products
    if not products:
        return redirect(url_for("index"))

    # Filter out previously known URLs
    previous_urls = set(_norm_url(u) for u in session.get("previous_urls", []))
    processed_urls = set()
    for p in session.get("accepted", []):
        processed_urls.add(_norm_url(p.get("product_url", "")))
    for url in session.get("rejected", []):
        processed_urls.add(_norm_url(url))

    # Brand filter from query param
    brand_filter = request.args.get("brand", "")

    # Only show products from ACTIVE brands (prevents ghost brands from old cache)
    active_brands = load_active_brands(country)
    active_brand_names = set(b["name"] for b in active_brands) if active_brands else set()

    # Single pass: compute remaining (filtered) + all_remaining (unfiltered) + brand counts
    remaining = []
    all_remaining = []
    for p in products:
        url = _norm_url(p["product_url"])
        if active_brand_names and p.get("brand", "") not in active_brand_names:
            continue
        if url not in previous_urls and url not in processed_urls:
            all_remaining.append(p)
            if not brand_filter or p.get("brand", "") == brand_filter:
                remaining.append(p)

    # Sort remaining by brand so products are grouped
    remaining.sort(key=lambda p: p.get("brand", ""))

    available_brands = {}
    for p in all_remaining:
        b = p.get("brand", "Desconocida")
        available_brands[b] = available_brands.get(b, 0) + 1
    sorted_brands = sorted(available_brands.items(), key=lambda x: -x[1])

    if not remaining:
        if brand_filter:
            # No more in this brand, but maybe others
            return redirect(url_for("curate"))
        return render_template("done.html",
                               accepted_count=len(session.get("accepted", [])),
                               total=len(products))

    current = remaining[0]
    current["_previous_count"] = len(session.get("previous_rows", []))
    progress = len(products) - len(all_remaining)

    return render_template("curate.html",
                           product=current,
                           remaining=len(remaining),
                           progress=progress,
                           total=len(products),
                           accepted_count=len(session.get("accepted", [])),
                           rejected_count=len(session.get("rejected", [])),
                           available_brands=sorted_brands,
                           current_brand_filter=brand_filter)


@app.route("/curate/next")
@login_required
def curate_next():
    """AJAX endpoint — returns BATCH of next 5 products for instant client-side switching"""
    country = load_active_country()
    session = load_session()

    # Use in-memory cache to avoid re-reading JSON file every request
    products = _get_cached_products(country)
    if not products:
        return jsonify({"done": True, "accepted": len(session.get("accepted", []))})

    previous_urls = set(_norm_url(u) for u in session.get("previous_urls", []))
    processed_urls = set()
    for p in session.get("accepted", []):
        processed_urls.add(_norm_url(p.get("product_url", "")))
    for url in session.get("rejected", []):
        processed_urls.add(_norm_url(url))

    active_brands = load_active_brands(country)
    active_brand_names = set(b["name"] for b in active_brands) if active_brands else set()
    brand_filter = request.args.get("brand", "")

    remaining = []
    all_remaining = []
    for p in products:
        url = _norm_url(p["product_url"])
        if active_brand_names and p.get("brand", "") not in active_brand_names:
            continue
        if url not in previous_urls and url not in processed_urls:
            all_remaining.append(p)
            if not brand_filter or p.get("brand", "") == brand_filter:
                remaining.append(p)

    available_brands = {}
    for p in all_remaining:
        b = p.get("brand", "Desconocida")
        available_brands[b] = available_brands.get(b, 0) + 1

    # Sort remaining by brand name so products are grouped
    remaining.sort(key=lambda p: p.get("brand", ""))

    if not remaining:
        return jsonify({"done": True, "accepted": len(session.get("accepted", [])),
                        "total": len(products)})

    batch_size = min(20, len(remaining))
    batch = remaining[:batch_size]
    prev_count = len(session.get("previous_rows", []))
    for p in batch:
        p["_previous_count"] = prev_count
    return jsonify({
        "done": False,
        "products": batch,          # Array of up to 5 products
        "product": batch[0],        # First product (backward compat)
        "remaining": len(remaining),
        "total": len(products),
        "progress": len(products) - len(all_remaining),
        "accepted_count": len(session.get("accepted", [])),
        "rejected_count": len(session.get("rejected", [])),
        "brands": sorted(available_brands.items(), key=lambda x: -x[1])
    })


@app.route("/undo", methods=["POST"])
@login_required
def undo_action():
    """Undo the last accept/reject/trend action."""
    data = request.get_json(silent=True) or {}
    last_action = data.get("action", "")
    product_url = data.get("product_url", "")
    if not product_url:
        return jsonify({"error": "No product to undo"}), 400

    session = load_session()
    norm = _norm_url(product_url)
    if last_action in ("accept", "trend"):
        session["accepted"] = [p for p in session["accepted"]
                               if _norm_url(p.get("product_url", "")) != norm]
    elif last_action == "reject":
        session["rejected"] = [u for u in session["rejected"]
                               if _norm_url(u) != norm]
    save_session(session)
    return jsonify({"status": "ok", "accepted": len(session["accepted"]),
                    "rejected": len(session["rejected"])})


@app.route("/action", methods=["POST"])
@login_required
def action():
    """Handle accept/reject/finish"""
    data = request.get_json(silent=True) or {}
    act = data.get("action")
    product = data.get("product")
    session = load_session()

    if act == "accept" and product:
        # Remove from rejected if it was there (dual-tab protection)
        url = _norm_url(product.get("product_url", ""))
        session["rejected"] = [u for u in session["rejected"] if _norm_url(u) != url]
        # Store only fields needed for spreadsheet (saves ~80% memory)
        slim = {k: product.get(k, "") for k in ("product_url", "brand", "name", "category", "tags", "price")}
        session["accepted"].append(slim)
    elif act == "trend" and product:
        url = _norm_url(product.get("product_url", ""))
        session["rejected"] = [u for u in session["rejected"] if _norm_url(u) != url]
        slim = {k: product.get(k, "") for k in ("product_url", "brand", "name", "category", "tags", "price")}
        slim["trend"] = True
        session["accepted"].append(slim)
    elif act == "reject" and product:
        url = _norm_url(product.get("product_url", ""))
        # Remove from accepted if it was there (dual-tab protection)
        session["accepted"] = [p for p in session["accepted"] if _norm_url(p.get("product_url", "")) != url]
        session["rejected"].append(url)
    elif act == "skip_brand":
        # Reject ALL remaining products from this brand
        brand_to_skip = data.get("brand", "")
        if brand_to_skip:
            country = load_active_country()
            cache_file = get_cache_file_for_country(country)
            products = []
            if os.path.exists(cache_file):
                with open(cache_file) as f:
                    products = json.load(f).get("products", [])
            if not products:
                products = load_cache_firestore(country) or []
            previous_urls = set(_norm_url(u) for u in session.get("previous_urls", []))
            processed = set(p.get("product_url", "").rstrip("/") for p in session.get("accepted", []))
            processed.update(u.rstrip("/") for u in session.get("rejected", []))
            for p in products:
                url = _norm_url(p["product_url"])
                if p.get("brand") == brand_to_skip and url not in previous_urls and url not in processed:
                    session["rejected"].append(url)
            save_session(session)
            return jsonify({"status": "ok", "skipped_brand": brand_to_skip})
    elif act == "finish":
        uid = get_user_id()
        country = load_active_country()
        output_path = os.path.join(DATA_DIR, f"moder_plantilla_{uid}.xlsx")
        previous_rows = session.get("previous_rows", [])

        # Load ALL crawled products for the spreadsheet
        all_products = _get_cached_products(country) or []
        if not all_products:
            return jsonify({"status": "error", "error": "No hay productos crawleados."}), 400

        # Build sets of approved and trend URLs
        accepted_urls = set()
        trend_urls = set()
        for p in session.get("accepted", []):
            url = _norm_url(p.get("product_url", ""))
            accepted_urls.add(url)
            if p.get("trend"):
                trend_urls.add(url)

        print(f"📊 FINISH: user={uid}, total={len(all_products)}, approved={len(accepted_urls)}, trends={len(trend_urls)}, previous={len(previous_rows)}")
        # brand_order: PRIORIDAD al override que el usuario eligió en el Paso 4
        # (guardado en session["brand_order_override"] por /save_order). Si no
        # hay override (user descargó directo sin pasar por Paso 4), fallback a
        # la selección original de active_brands.
        #
        # Fix (11 abril 2026) — antes esta línea SIEMPRE leía active_brands y
        # ignoraba el override, causando que el Excel generado desde el botón
        # "Descargar planilla" del Paso 4 respetara Posición pero NO Orden.
        _brand_order_from_selection = session.get("brand_order_override") or [
            b.get("name", "") for b in (load_active_brands(country) or []) if b.get("name")
        ]
        _product_order_override = session.get("product_order_override") or None
        try:
            _, xlsx_buffer = generate_plantilla(
                all_products, accepted_urls, trend_urls, output_path,
                previous_rows=previous_rows,
                brand_order=_brand_order_from_selection,
                product_order_override=_product_order_override,
            )
            with _global_cache_lock:
                if len(generated_xlsx_per_user) >= _XLSX_CACHE_MAX:
                    oldest = next(iter(generated_xlsx_per_user))
                    del generated_xlsx_per_user[oldest]
                generated_xlsx_per_user[uid] = xlsx_buffer
            print(f"✅ Planilla generada: {len(all_products)} productos → {output_path}")
        except Exception as e:
            print(f"Error generating spreadsheet: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({"status": "error", "error": str(e)}), 500
        total = len(previous_rows) + len(session["accepted"])
        accepted_count = len(session["accepted"])
        # DO NOT clear session here — keep accepted until download completes
        # Session will be cleared when user starts a new crawl or cancels
        save_session(session)
        return jsonify({"status": "done", "count": accepted_count,
                        "previous": len(previous_rows), "total": total, "path": output_path})

    save_session(session)
    return jsonify({"status": "ok", "accepted": len(session["accepted"])})


@app.route("/download")
@login_required
def download():
    """Download generated spreadsheet per user"""
    uid = get_user_id()
    filename = f"moder_plantilla_{uid}.xlsx"
    path = os.path.join(DATA_DIR, filename)

    # Try disk file first (per-user)
    if os.path.exists(path):
        try:
            return send_file(path, as_attachment=True, download_name="moder_plantilla_productos.xlsx")
        except Exception as e:
            print(f"Error sending file from disk: {e}")

    # Fallback: serve from per-user in-memory buffer
    if uid in generated_xlsx_per_user and generated_xlsx_per_user[uid] is not None:
        try:
            generated_xlsx_per_user[uid].seek(0)
            return send_file(
                generated_xlsx_per_user[uid],
                as_attachment=True,
                download_name="moder_plantilla_productos.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            print(f"Error sending file from memory: {e}")

    # Last resort: regenerate from user's session data
    session = load_session()
    country = load_active_country()
    all_products = _get_cached_products(country) or []
    if all_products:
        try:
            previous_rows = session.get("previous_rows", [])
            accepted_urls = set(_norm_url(p.get("product_url", "")) for p in session.get("accepted", []))
            trend_urls = set(_norm_url(p.get("product_url", "")) for p in session.get("accepted", []) if p.get("trend"))
            # Fix (11 abril 2026) — respetar brand_order_override del Paso 4
            _brand_order_from_selection = session.get("brand_order_override") or [
                b.get("name", "") for b in (load_active_brands(country) or []) if b.get("name")
            ]
            _product_order_override = session.get("product_order_override") or None
            _, xlsx_buffer = generate_plantilla(
                all_products, accepted_urls, trend_urls, path,
                previous_rows=previous_rows,
                brand_order=_brand_order_from_selection,
                product_order_override=_product_order_override,
            )
            with _global_cache_lock:
                if len(generated_xlsx_per_user) >= _XLSX_CACHE_MAX:
                    oldest = next(iter(generated_xlsx_per_user))
                    del generated_xlsx_per_user[oldest]
                generated_xlsx_per_user[uid] = xlsx_buffer
            xlsx_buffer.seek(0)
            return send_file(
                xlsx_buffer,
                as_attachment=True,
                download_name="moder_plantilla_productos.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            print(f"Error regenerating spreadsheet: {e}")
            import traceback
            traceback.print_exc()
            import html as html_mod
            return f"<html><body><h2>Error al generar planilla</h2><p>{html_mod.escape(str(e))}</p><a href='/'>Volver</a></body></html>", 500

    return "<html><body><h2>No se ha generado la planilla aún</h2><p>Primero acepta productos y presiona Finalizar.</p><a href='/'>Volver al inicio</a></body></html>", 404


# ─────────────────────────────────────────────────────────────────────────────
# PASO 4: "Ordenar y Descargar" — visualizador con flechas ↑↓ para reordenar
# marcas y productos antes de generar la planilla o subir al admin.
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/order")
@login_required
def order_view():
    """Muestra la UI del Paso 4: reordenar marcas aprobadas y sus productos."""
    session = load_session()
    country = load_active_country()

    # Determinar el orden inicial de marcas: preferir un override previo que el
    # user haya guardado antes (si entra y sale de /order sin descargar), o el
    # orden de selección de active_brands como fallback inicial.
    prior_brand_order = session.get("brand_order_override")
    if not prior_brand_order:
        prior_brand_order = [b.get("name", "") for b in (load_active_brands(country) or []) if b.get("name")]

    brands = build_curated_brands_for_ordering(
        session, country, brand_selection_order=prior_brand_order
    )

    # Si hay un product_order_override previo, aplicarlo dentro de cada marca
    product_order_override = session.get("product_order_override") or {}
    if product_order_override:
        for b in brands:
            desired = product_order_override.get(b["name"])
            if not desired:
                continue
            # Re-ordenar preservando el índice en `desired`. Productos nuevos
            # (no vistos en el override anterior) van al final.
            def _rank(p, order_list=desired):
                try:
                    return order_list.index(_norm_url(p["url"]))
                except ValueError:
                    return len(order_list) + 9999
            b["products"].sort(key=_rank)

    # Si también hay overrides de tags por producto, aplicarlos
    tag_overrides = session.get("product_tag_overrides") or {}
    for b in brands:
        for p in b["products"]:
            key = _norm_url(p["url"])
            if key in tag_overrides:
                p["tags"] = tag_overrides[key]

    total_products = sum(len(b["products"]) for b in brands)
    return render_template(
        "order.html",
        brands=brands,
        brand_count=len(brands),
        total_products=total_products,
    )


@app.route("/save_order", methods=["POST"])
@login_required
def save_order():
    """Persiste el orden visual del Paso 4 en la session.

    Payload JSON esperado:
      {
        "brand_order": ["Marca A", "Marca B", ...],
        "product_order": {
            "Marca A": ["https://url1", "https://url2", ...],
            "Marca B": [...]
        },
        "tags": {
            "https://url1": ["etiqueta1", "etiqueta2", "etiqueta3", "etiqueta4"],
            ...
        }
      }

    Estos overrides se leen por `generate_plantilla` y por `upload_to_admin`
    cuando se construye la salida final. No se mezclan con los datos de crawl
    originales; son una "capa de curación" que vive solo en la session.
    """
    data = request.get_json(silent=True) or {}
    brand_order = data.get("brand_order") or []
    product_order = data.get("product_order") or {}
    tags = data.get("tags") or {}

    if not isinstance(brand_order, list):
        return jsonify({"status": "error", "error": "brand_order debe ser lista"}), 400
    if not isinstance(product_order, dict):
        return jsonify({"status": "error", "error": "product_order debe ser dict"}), 400
    if not isinstance(tags, dict):
        return jsonify({"status": "error", "error": "tags debe ser dict"}), 400

    # Normalizar URLs para que la clave sea idéntica a la usada en
    # generate_plantilla (que compara con _norm_url).
    normalized_product_order = {}
    for brand, urls in product_order.items():
        if not isinstance(urls, list):
            continue
        normalized_product_order[brand] = [_norm_url(u) for u in urls if isinstance(u, str)]

    normalized_tags = {}
    for url, tag_list in tags.items():
        if not isinstance(tag_list, list):
            continue
        # Máximo 4 tags, trim, descartar vacíos
        cleaned = [str(t).strip() for t in tag_list[:4]]
        normalized_tags[_norm_url(url)] = cleaned

    session = load_session()
    session["brand_order_override"] = [str(b) for b in brand_order]
    session["product_order_override"] = normalized_product_order
    session["product_tag_overrides"] = normalized_tags
    save_session(session)

    return jsonify({
        "status": "ok",
        "brands": len(brand_order),
        "product_brands": len(normalized_product_order),
        "tagged_products": len(normalized_tags),
    })


@app.route("/upload_to_admin", methods=["POST"])
@login_required
def upload_to_admin():
    """Escribe las marcas curadas directamente a Firestore (collection `stores`)
    usando el SDK que el robot ya tiene cargado via firestore_storage.py.

    Este endpoint respeta los overrides del Paso 4 (brand order, product order,
    tags) si existen en la session. Si no, usa el orden natural de crawl.

    Merge policy: REEMPLAZO COMPLETO por store_id. Si `stores/{id}` ya existe
    en Firestore, se sobrescribe con el nuevo payload. Confirmado con el user
    en el checkpoint del 11 abril 2026.
    """
    session = load_session()
    country = load_active_country()
    uid = get_user_id()

    # Brand order: prefer override from Paso 4, fallback a active_brands.
    brand_order = session.get("brand_order_override") or [
        b.get("name", "") for b in (load_active_brands(country) or []) if b.get("name")
    ]

    brands = build_curated_brands_for_ordering(
        session, country, brand_selection_order=brand_order
    )

    if not brands:
        return jsonify({
            "status": "error",
            "error": "No hay marcas curadas para subir. Aprueba productos primero."
        }), 400

    # Apply product order override within each brand
    product_order_override = session.get("product_order_override") or {}
    for b in brands:
        desired = product_order_override.get(b["name"])
        if not desired:
            continue
        def _rank(p, order_list=desired):
            try:
                return order_list.index(_norm_url(p["url"]))
            except ValueError:
                return len(order_list) + 9999
        b["products"].sort(key=_rank)

    # Apply tag overrides
    tag_overrides = session.get("product_tag_overrides") or {}
    for b in brands:
        for p in b["products"]:
            key = _norm_url(p["url"])
            if key in tag_overrides:
                p["tags"] = tag_overrides[key]

    # Firestore write via firestore_storage helper (already imported).
    # Uso `_get_db` porque es el handle único que la app ya usa para persistir
    # session/brands/cache. Si no está inicializado (por ejemplo porque
    # FIREBASE_SERVICE_ACCOUNT no está en el env), devuelve None y retornamos
    # un error explícito para que el admin use el flujo manual de Excel.
    try:
        from firestore_storage import _get_db
    except ImportError:
        return jsonify({
            "status": "error",
            "error": "firestore_storage no disponible en esta instancia del robot."
        }), 500

    db = _get_db()
    if db is None:
        return jsonify({
            "status": "error",
            "error": "Firestore no configurado en el robot. Usa el botón 'Descargar planilla' y súbela manualmente al admin."
        }), 500

    # Build and write each store. batch para atomicidad.
    from google.cloud.firestore_v1 import SERVER_TIMESTAMP  # type: ignore
    batch = db.batch()
    now_ts = SERVER_TIMESTAMP
    stores_collection = db.collection("stores")

    # ── Merge policy: REEMPLAZO COMPLETO DE COLECCIÓN POR PAÍS ──
    # Fix (12 abril 2026) — el comportamiento anterior hacía batch.set por
    # store_id, lo cual sobrescribía los stores del upload actual pero dejaba
    # intactas las tiendas viejas de uploads anteriores. Eso generaba un
    # merge-implícito al nivel de colección.
    #
    # Ahora: borramos TODAS las tiendas del país activo antes de escribir las
    # nuevas. Resultado: al final del batch commit, Firestore tiene SOLO las
    # tiendas del curado actual, nada más.
    #
    # Nota de seguridad: el delete + set van en el MISMO batch, por lo que
    # todo es atómico. Si el batch falla, no borramos nada.
    stores_to_delete = []
    try:
        existing_query = stores_collection.where("country", "==", country).stream()
        for existing_doc in existing_query:
            stores_to_delete.append(existing_doc.id)
    except Exception as _list_err:
        # Si el where falla (ej: sin índice compuesto, o country no estaba
        # indexado), caemos a un listado sin filtro y filtramos en cliente.
        # Esto es más lento pero robusto.
        print(f"⚠️ country filter falló, usando scan full: {_list_err}")
        for existing_doc in stores_collection.stream():
            data = existing_doc.to_dict() or {}
            doc_country = data.get("country", "")
            # Si el doc no tiene country explícito, inferir del id (termina en _xx)
            if not doc_country and existing_doc.id.endswith(f"_{country.lower()}"):
                doc_country = country
            if doc_country == country:
                stores_to_delete.append(existing_doc.id)

    for old_id in stores_to_delete:
        batch.delete(stores_collection.document(old_id))

    print(f"🗑️  [upload_to_admin] Borrando {len(stores_to_delete)} tiendas existentes del país {country}")

    written_stores = []
    total_products = 0
    for brand_idx, b in enumerate(brands, start=1):
        brand_name = b["name"]
        # store_id estable: slug de marca + país (e.g. "manto_silvestre_cl")
        slug = "".join(c if c.isalnum() else "_" for c in brand_name.lower()).strip("_")
        store_id = f"{slug}_{country.lower()}"

        products_payload = []
        for pos_in_brand, p in enumerate(b["products"], start=1):
            tags_clean = [str(t).strip() for t in (p.get("tags") or []) if str(t).strip()]
            # Pad tags a 4 posiciones para un schema consistente
            while len(tags_clean) < 4:
                tags_clean.append("")

            products_payload.append({
                "id": f"{store_id}_p{pos_in_brand}",
                "url": p.get("url", ""),
                "imageURL": p.get("image_url", ""),
                "title": p.get("title", ""),
                "order": pos_in_brand,          # orden del producto dentro del store
                "country": country,
                "isTop20": False,
                "isTrend": bool(p.get("is_trend")),
                "category": p.get("category", ""),
                "tag1": tags_clean[0],
                "tag2": tags_clean[1],
                "tag3": tags_clean[2],
                "tag4": tags_clean[3],
                "tags": [t for t in tags_clean if t],
            })

        store_doc = {
            "name": brand_name,
            "products": products_payload,
            "isActive": True,
            "order": brand_idx,                 # rank de la marca (importancia)
            "country": country,
            "updatedAt": now_ts,
            "updatedBy": f"robot_{uid}",
        }

        doc_ref = stores_collection.document(store_id)
        # Merge policy: reemplazo completo del doc (set sin merge).
        # Si existe, se sobrescribe. Si no, se crea.
        batch.set(doc_ref, store_doc)
        written_stores.append(store_id)
        total_products += len(products_payload)

    try:
        batch.commit()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            "status": "error",
            "error": f"Firestore batch commit falló: {str(e)[:200]}"
        }), 500

    return jsonify({
        "status": "ok",
        "stores": len(written_stores),
        "products": total_products,
        "store_ids": written_stores,
    })


@app.route("/cancel-curation", methods=["POST"])
@login_required
def cancel_curation():
    """Cancel active curation — clears session, cache, and stops crawl"""
    uid = get_user_id()
    country = load_active_country()
    clear_curation_session(country)
    # Signal crawl thread to stop (if running)
    crawl_cancel_event.set()
    # Don't clear the event here — let the crawl thread clear it when it exits
    # This ensures the crawl sees the cancel signal even if it's sleeping between brands
    uid = get_user_id()
    for k in list(crawl_progress.keys()):
        if k not in _default_progress:
            del crawl_progress[k]
    crawl_progress.update(_default_progress)
    return jsonify({"status": "ok", "message": "Curación cancelada. Sesión limpia."})


def clear_curation_session(country=None, user_id=None):
    """Clear all curation data — session, cache, crawl state (local + Firestore)"""
    uid = user_id or get_user_id()
    # Clear in-memory cache
    invalidate_session_cache(uid)
    # Clear local per-user session file
    user_session = _session_file_for_user(uid)
    if os.path.exists(user_session):
        os.remove(user_session)
    # Also clean legacy session file
    if os.path.exists(SESSION_FILE):
        os.remove(SESSION_FILE)
    if country:
        cache = get_cache_file_for_country(country, uid)
        if os.path.exists(cache):
            os.remove(cache)
    if os.path.exists(CRAWL_CACHE):
        os.remove(CRAWL_CACHE)
    # Clear Firestore
    clear_session_firestore(uid)
    if country:
        clear_cache_firestore(country)
    # Reset progress per user
    for k in list(crawl_progress.keys()):
        if k not in _default_progress:
            del crawl_progress[k]
    crawl_progress.update(_default_progress)


@app.route("/reset", methods=["POST"])
@login_required
def reset():
    """Reset everything — session, cache, brands, country — complete clean slate"""
    uid = get_user_id()
    country = load_active_country()
    clear_curation_session(country, uid)
    # Clear per-user brands file + Firestore
    if country:
        brands_path = _brands_file_for_user(country, uid)
        if os.path.exists(brands_path):
            os.remove(brands_path)
        save_brands_firestore([], f"{uid}_{country}")
    # Clear legacy files
    for f in [BRANDS_FILE,
              os.path.join(DATA_DIR, f"previous_upload_{uid}.xlsx"),
              os.path.join(DATA_DIR, f"moder_plantilla_{uid}.xlsx")]:
        if os.path.exists(f):
            os.remove(f)
    # Clear all country brand files for this user
    import glob
    for f in glob.glob(os.path.join(DATA_DIR, f"brands_{uid}_*.json")):
        os.remove(f)
    # Clear country
    save_active_country("")
    return jsonify({"status": "ok"})


# ─── Run ─────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    os.makedirs(os.path.join(DATA_DIR, "templates"), exist_ok=True)
    print("\n🎨 MODÈR Product Curator v2")
    print("   http://localhost:5050\n")
    app.run(host="0.0.0.0", port=5050, debug=False)
